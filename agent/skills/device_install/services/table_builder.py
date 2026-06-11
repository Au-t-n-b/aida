"""
table_builder · 生成三类 xlsx 产物。

- generate_principal_table：责任人信息表（按真实任务行 + 已填责任人/责任主体生成）。
- generate_full_tasks_xlsx：设备安装全量任务表。
- generate_dispatch_plan_xlsx：设备安装实施计划（按管理单元分组）。
"""
from __future__ import annotations

import os


def _styles():
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    return {
        "header_fill": PatternFill("solid", fgColor="1F4E79"),
        "header_font": Font(color="FFFFFF", bold=True, name="微软雅黑"),
        "alt_fill": PatternFill("solid", fgColor="EBF3FB"),
        "group_fill": PatternFill("solid", fgColor="D6E4F0"),
        "thin": Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        ),
        "Alignment": Alignment,
    }


def _write_header(ws, headers: list[str], st: dict) -> None:
    Alignment = st["Alignment"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = st["header_fill"]
        c.font = st["header_font"]
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = st["thin"]
    ws.row_dimensions[1].height = 22


def _set_widths(ws, widths: list[int]) -> None:
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[chr(ord("A") + ci - 1)].width = w


def generate_principal_table(tasks: list[dict], output_path: str) -> bool:
    """根据任务列表生成《责任人信息表.xlsx》。

    每行对应一条真实安装任务（来自《任务计划表》解析），责任人/责任主体取自
    已合并的真实值（数据中心《责任人信息表》或用户在线填写），不写任何 mock 数据。
    """
    try:
        import openpyxl  # type: ignore
    except ImportError:
        return False
    st = _styles()
    Alignment = st["Alignment"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "责任人信息"

    headers = [
        "序列号", "管理单元", "活动ID", "活动名称",
        "开始日期", "结束日期", "责任人姓名", "责任主体（华为/分包商）", "备注",
    ]
    _write_header(ws, headers, st)
    for ri, t in enumerate(tasks, 2):
        row_vals = [
            t.get("id", ""), t.get("unit", ""), t.get("activity_id", ""),
            t.get("activity_name", ""), t.get("start_date", ""), t.get("end_date", ""),
            t.get("principal") or t.get("owner", ""),
            t.get("principal_org", ""),
            "",
        ]
        for ci, v in enumerate(row_vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = st["thin"]
            c.alignment = Alignment(vertical="center")
            if ri % 2 == 0:
                c.fill = st["alt_fill"]
    _set_widths(ws, [14, 16, 10, 28, 14, 14, 16, 24, 20])

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return True


def generate_full_tasks_xlsx(tasks: list[dict], output_path: str) -> bool:
    """生成《设备安装全量任务.xlsx》。"""
    try:
        import openpyxl  # type: ignore
    except ImportError:
        return False
    st = _styles()
    Alignment = st["Alignment"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "设备安装全量任务"

    headers = [
        "序号", "管理单元", "活动ID", "活动名称",
        "开始日期", "结束日期", "SLA", "设备型号&数量",
        "依赖活动", "责任人", "责任主体", "远程团队", "现场团队", "状态",
    ]
    _write_header(ws, headers, st)
    for ri, t in enumerate(tasks, 2):
        row_vals = [
            ri - 1,
            t.get("unit", ""), t.get("activity_id", ""),
            t.get("activity_name", ""), t.get("start_date", ""),
            t.get("end_date", ""), t.get("sla", ""),
            t.get("devices", ""), t.get("dependencies", ""),
            t.get("principal") or t.get("owner", ""),
            t.get("principal_org", ""),
            t.get("remote_team", ""), t.get("local_team", ""),
            t.get("status", "待下发"),
        ]
        for ci, v in enumerate(row_vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = st["thin"]
            c.alignment = Alignment(vertical="center")
            if ri % 2 == 0:
                c.fill = st["alt_fill"]
    _set_widths(ws, [6, 16, 8, 28, 14, 14, 8, 40, 24, 16, 16, 12, 12, 10])

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return True


def generate_dispatch_plan_with_sn_xlsx(
    tasks: list[dict],
    position_path: str,
    arrival_path: str,
    output_path: str,
) -> bool:
    """生成**自包含**《设备安装实施计划.xlsx》（双 Sheet）。

    Sheet「设备安装实施计划」：现有实施计划列 + 首列「计划行ID」(= 管理单元::活动ID)，
        按管理单元分组，全量任务。
    Sheet「SN扫码表」：首列「所属管理单元」(外键回指实施计划) + 现有 SN 列，ESN 留空；
        内容由 build_sn_rows_for_plan 按全量任务配额匹配位置表得到。

    解析侧：Sheet1 → tasks_state；Sheet2 按 (所属机房, 设备大类) 分组即还原 sn_tables 结构。
    """
    try:
        import openpyxl  # type: ignore
        from openpyxl.styles import PatternFill
    except ImportError:
        return False
    from .sn_builder import build_sn_rows_for_plan
    from ._common import as_str

    st = _styles()
    Alignment = st["Alignment"]

    wb = openpyxl.Workbook()

    # —— Sheet 1：设备安装实施计划 ——
    ws = wb.active
    ws.title = "设备安装实施计划"
    headers = [
        "计划行ID", "管理单元", "活动ID", "活动名称",
        "计划开始", "计划完成", "SLA", "责任人", "责任主体", "设备型号&数量", "状态",
    ]
    _write_header(ws, headers, st)

    by_unit: dict[str, list[dict]] = {}
    for t in tasks:
        by_unit.setdefault(t.get("unit", "未知"), []).append(t)

    row_idx = 2
    for unit, unit_tasks in sorted(by_unit.items()):
        for t in unit_tasks:
            plan_id = f"{as_str(t.get('unit'))}::{as_str(t.get('activity_id'))}"
            row_vals = [
                plan_id, unit,
                t.get("activity_id", ""), t.get("activity_name", ""),
                t.get("start_date", ""), t.get("end_date", ""),
                t.get("sla", ""),
                t.get("principal") or t.get("owner", ""),
                t.get("principal_org", ""),
                t.get("devices", ""),
                t.get("status", "待下发"),
            ]
            for ci, v in enumerate(row_vals, 1):
                c = ws.cell(row=row_idx, column=ci, value=v)
                c.border = st["thin"]
                c.alignment = Alignment(vertical="center", wrap_text=True)
                c.fill = st["group_fill"]
            row_idx += 1
    _set_widths(ws, [18, 16, 8, 28, 14, 14, 8, 16, 16, 40, 10])

    # —— Sheet 2：SN扫码表（全量 · 带所属管理单元 + 关联计划行ID）——
    # 关联计划行ID（= 管理单元::活动ID，多活动用 ; 分隔）建立 SN ⇄ 实施计划任务级对应，
    # 下游按「计划下发」勾选的计划行ID 过滤设备，实现任务级 ESN 录入。
    ws2 = wb.create_sheet("SN扫码表")
    sn_headers = [
        "所属管理单元", "关联计划行ID", "设备大类", "厂家", "设备型号", "设备名称",
        "所属机房", "所属机柜", "安装起始U位", "设备U高", "ESN（待填）",
    ]
    _write_header(ws2, sn_headers, st)
    sn_keys = [
        "所属管理单元", "关联计划行ID", "设备大类", "厂家", "设备型号", "设备名称",
        "所属机房", "所属机柜", "安装起始U位", "设备U高", "ESN",
    ]
    esn_fill = PatternFill("solid", fgColor="FFF2CC")
    sn_rows = build_sn_rows_for_plan(position_path, arrival_path, tasks)
    for ri, r in enumerate(sn_rows, 2):
        for ci, key in enumerate(sn_keys, 1):
            c = ws2.cell(row=ri, column=ci, value=r.get(key, ""))
            c.border = st["thin"]
            c.alignment = Alignment(vertical="center")
            if key == "ESN":
                c.fill = esn_fill
    _set_widths(ws2, [16, 22, 14, 10, 22, 40, 10, 10, 10, 8, 28])
    ws2.row_dimensions[1].height = 22

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return True


def generate_dispatch_plan_xlsx(tasks: list[dict], output_path: str) -> bool:
    """生成《设备安装实施计划.xlsx》，按管理单元分组展示。"""
    try:
        import openpyxl  # type: ignore
    except ImportError:
        return False
    st = _styles()
    Alignment = st["Alignment"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "设备安装实施计划"

    headers = [
        "序号", "管理单元", "活动ID", "活动名称",
        "计划开始", "计划完成", "SLA", "责任人", "责任主体", "设备型号&数量", "状态",
    ]
    _write_header(ws, headers, st)

    by_unit: dict[str, list[dict]] = {}
    for t in tasks:
        by_unit.setdefault(t.get("unit", "未知"), []).append(t)

    row_idx, seq = 2, 1
    for unit, unit_tasks in sorted(by_unit.items()):
        for t in unit_tasks:
            row_vals = [
                seq, unit,
                t.get("activity_id", ""), t.get("activity_name", ""),
                t.get("start_date", ""), t.get("end_date", ""),
                t.get("sla", ""),
                t.get("principal") or t.get("owner", ""),
                t.get("principal_org", ""),
                t.get("devices", ""),
                t.get("status", "待下发"),
            ]
            for ci, v in enumerate(row_vals, 1):
                c = ws.cell(row=row_idx, column=ci, value=v)
                c.border = st["thin"]
                c.alignment = Alignment(vertical="center", wrap_text=True)
                c.fill = st["group_fill"]
            row_idx += 1
            seq += 1
    _set_widths(ws, [6, 16, 8, 28, 14, 14, 8, 16, 16, 40, 10])

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return True
