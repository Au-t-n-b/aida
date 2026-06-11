"""
sn_builder · SN 扫码表生成与回读。

- build_sn_tables_data：读《设备位置表》+《到货信息表》，按（机房, 设备大类）分组。
- generate_sn_xlsx：为一个分组生成 SN 扫码表 xlsx（ESN 列留空待填）。
- read_filled_sn_table：回读用户填好的 SN 扫码表，提取每行 ESN。
"""
from __future__ import annotations

import os
import re
from collections import defaultdict

from ._common import as_str

SN_COL_KEYS = ["设备大类", "厂家", "设备型号", "设备名称", "所属机房", "所属机柜", "安装起始U位", "设备U高", "ESN"]

# 任务「设备型号&数量列表」单元格拆分
_SPLIT_RE = re.compile(r"[、,，;；/\n\r]+")
# *6 / x6 / ×6 后缀
_QTY_STAR_RE = re.compile(r"^(.+?)[\*xX×]\s*(\d+)\s*$")
# 「CE6881-48S6CQ 2台」「Atlas 900 A3 SuperPoD计算柜 12柜」「400G MPO线缆 2688根」
_QTY_UNIT_RE = re.compile(r"^(.+?)\s+(\d+)\s*([台柜根个]+)\s*$")


def _norm_model(s: str) -> str:
    """归一化型号用于匹配：去空格、转小写。"""
    return as_str(s).replace(" ", "").lower()


def parse_device_entries(devices: str) -> list[tuple[str, int]]:
    """解析「设备型号&数量列表」→ [(型号规格, 数量), ...]。
    支持 *6 / 2台 / 12柜 / 2688根 等写法；无法解析数量时 qty=0。"""
    out: list[tuple[str, int]] = []
    for part in _SPLIT_RE.split(as_str(devices)):
        p = part.strip()
        if not p:
            continue
        m = _QTY_STAR_RE.match(p)
        if m:
            out.append((m.group(1).strip(), int(m.group(2))))
            continue
        m = _QTY_UNIT_RE.match(p)
        if m:
            out.append((m.group(1).strip(), int(m.group(2))))
            continue
        out.append((p, 0))
    return out


def parse_device_models(devices: str) -> set[str]:
    """从任务的「设备型号&数量列表」解析出型号集合（去数量后缀）。"""
    return {model for model, _ in parse_device_entries(devices) if model}


def aggregate_device_quotas(dispatched_tasks: list[dict]) -> dict[str, int]:
    """根据「计划下发」勾选任务汇总设备配额。

    - 同一管理单元 + 同一型号规格在多任务（安装/上电/验收）重复出现时取最大数量，避免重复累加。
    - 跨管理单元（如 101-PoD1 + 101-PoD2）对同一规格数量相加。
    返回 {normalized_任务侧型号规格: 待匹配台数}；线缆等无法匹配机架位号的规格自然被过滤掉。
    """
    peaks: dict[tuple[str, str], int] = {}
    for t in dispatched_tasks:
        unit = as_str(t.get("unit"))
        devices = as_str(t.get("devices"))
        if not devices:
            continue
        for model, qty in parse_device_entries(devices):
            spec = _norm_model(model)
            if not spec:
                continue
            k = (unit, spec)
            peaks[k] = max(peaks.get(k, 0), qty if qty > 0 else 0)

    totals: dict[str, int] = defaultdict(int)
    for (_unit, spec), qty in peaks.items():
        if qty > 0:
            totals[spec] += qty
    return dict(totals)


def _peak_unit_spec_quotas(tasks: list[dict]) -> dict[tuple[str, str], int]:
    """按 (管理单元, 归一化型号) 取峰值数量（安装/上电/验收 同单元同型号去重）。

    与 aggregate_device_quotas 的区别：**保留管理单元维度**（后者把 unit 汇总掉了），
    使每台匹配到的物理设备能回填「所属管理单元」，从而建立实施计划 ⇄ SN 的对应关系。
    """
    peaks: dict[tuple[str, str], int] = {}
    for t in tasks:
        unit = as_str(t.get("unit"))
        for model, qty in parse_device_entries(as_str(t.get("devices"))):
            spec = _norm_model(model)
            if not spec:
                continue
            k = (unit, spec)
            peaks[k] = max(peaks.get(k, 0), qty if qty > 0 else 0)
    return {k: v for k, v in peaks.items() if v > 0}


def unit_spec_plan_ids(tasks: list[dict]) -> dict[tuple[str, str], list[str]]:
    """{(管理单元, 归一化型号): [计划行ID,...]}；计划行ID = 管理单元::活动ID。

    建立「设备规格 → 任务」反查表：同一 (unit, spec) 被多活动（安装/上电/验收）
    引用时收集全部计划行ID，使每台物理设备能回指其全部关联任务（任务级 ESN 对齐）。
    """
    out: dict[tuple[str, str], list[str]] = {}
    for t in tasks:
        unit = as_str(t.get("unit"))
        aid = as_str(t.get("activity_id"))
        plan_id = as_str(t.get("plan_row_id")) or f"{unit}::{aid}"
        for model, _qty in parse_device_entries(as_str(t.get("devices"))):
            spec = _norm_model(model)
            if not spec:
                continue
            lst = out.setdefault((unit, spec), [])
            if plan_id and plan_id not in lst:
                lst.append(plan_id)
    return out


def build_sn_rows_for_plan(
    position_path: str,
    arrival_path: str,
    tasks: list[dict],
) -> list[dict]:
    """为**全量实施计划**构建扁平 SN 行，每行携带「所属管理单元」。

    与 build_sn_tables_data 的差异：
      - 不依赖「计划下发」勾选，按全量任务的设备配额匹配位置表（接收侧勾选前即可生成）；
      - 返回扁平行（非按机房分组），每行带 所属管理单元，作为合并文件 Sheet2 的内容；
        解析回来时再按 (所属机房, 设备大类) 分组即可还原 sn_tables.json 结构。

    返回：[{所属管理单元, 设备大类, 厂家, 设备型号, 设备名称,
            所属机房, 所属机柜, 安装起始U位, 设备U高, ESN:""}, ...]
    """
    try:
        import openpyxl  # type: ignore
    except ImportError:
        return []

    if not position_path or not os.path.isfile(position_path):
        return []

    quotas = _peak_unit_spec_quotas(tasks)  # (unit, spec) -> remaining
    plan_ids_map = unit_spec_plan_ids(tasks)  # (unit, spec) -> [计划行ID,...]

    model_to_type: dict[str, str] = {}
    if arrival_path and os.path.isfile(arrival_path):
        try:
            wb_a = openpyxl.load_workbook(arrival_path, read_only=True, data_only=True)
            ws_a = wb_a.active
            a_rows = list(ws_a.iter_rows(values_only=True))
            wb_a.close()
            for row in a_rows[1:]:
                if len(row) > 2 and row[1] and row[2]:
                    model_to_type[as_str(row[2])] = as_str(row[1])
        except Exception:
            pass

    try:
        wb_p = openpyxl.load_workbook(position_path, read_only=True, data_only=True)
        sheet_name = "设备位置信息" if "设备位置信息" in wb_p.sheetnames else wb_p.sheetnames[-1]
        ws_p = wb_p[sheet_name]
        pos_rows = list(ws_p.iter_rows(values_only=True))
        wb_p.close()
    except Exception:
        return []

    rows_out: list[dict] = []
    # 稳定遍历配额键，保证同一输入产出一致
    quota_keys = sorted(quotas.keys())
    for row in pos_rows[1:]:
        if not row or not row[0]:
            continue
        model = as_str(row[0])
        matched_key: tuple[str, str] | None = None
        for key in quota_keys:
            if quotas.get(key, 0) <= 0:
                continue
            _unit, spec = key
            if _model_allowed(model, {spec}):
                matched_key = key
                break
        if matched_key is None:
            continue
        quotas[matched_key] -= 1

        unit = matched_key[0]
        dev_name  = as_str(row[2]) if len(row) > 2 else ""
        room      = as_str(row[3]) if len(row) > 3 else ""
        col       = as_str(row[4]) if len(row) > 4 else ""
        u_pos     = as_str(row[5]) if len(row) > 5 else ""
        u_height  = as_str(row[6]) if len(row) > 6 else ""
        dev_class = get_sn_device_class(model, model_to_type)
        rows_out.append({
            "所属管理单元": unit,
            "关联计划行ID": ";".join(plan_ids_map.get(matched_key, [])),
            "设备大类":    dev_class,
            "厂家":       get_sn_vendor(model),
            "设备型号":   model,
            "设备名称":   dev_name,
            "所属机房":   room,
            "所属机柜":   col,
            "安装起始U位": u_pos,
            "设备U高":    u_height,
            "ESN":        "",
        })
    return rows_out


def _model_allowed(model: str, allowed_norm: set[str]) -> bool:
    """型号是否落在允许集合（归一化后双向包含匹配，容忍命名细微差异）。"""
    nm = _norm_model(model)
    if not nm:
        return False
    return any(nm == a or nm in a or a in nm for a in allowed_norm)


def get_sn_device_class(model: str, model_to_type: dict[str, str]) -> str:
    """根据设备型号推断设备大类（先精确匹配到货表，再按型号关键字兜底）。"""
    m = model.strip()
    if m in model_to_type:
        return model_to_type[m]
    for k, v in model_to_type.items():
        if m in k or k in m:
            return v
    if "Atlas" in m:
        return "计算柜"
    if "LingQu" in m or "LQ" in m:
        return "总线设备柜"
    if "OceanStor" in m or "A800" in m:
        return "存储"
    if any(x in m for x in ["CE", "XH", "S5735", "LEAF", "SPINE"]):
        return "网络"
    return "其他"


def get_sn_vendor(model: str) -> str:
    huawei_keywords = ["Atlas", "CE", "XH", "OceanStor", "S5735", "LingQu", "A800", "CloudEngine"]
    return "华为" if any(k in model for k in huawei_keywords) else ""


def build_sn_tables_data(
    position_path: str,
    arrival_path: str = "",
    *,
    dispatched_tasks: list[dict] | None = None,
) -> list[dict]:
    """读《设备位置表》(+《到货信息表》)，按（机房, 设备大类）分组返回 SN 扫码表元数据。
    返回：[{id, room, device_class, file, rows:[...]}]

    dispatched_tasks 非空时，仅纳入「计划下发」勾选任务涉及的设备：
    按任务侧型号规格与位置表型号模糊匹配，并按任务数量配额截取（非全量位置表）。
    """
    try:
        import openpyxl  # type: ignore
    except ImportError:
        return []

    if not position_path or not os.path.isfile(position_path):
        return []

    quotas: dict[str, int] = (
        aggregate_device_quotas(dispatched_tasks) if dispatched_tasks else {}
    )

    # 到货信息表 → 型号→设备大类 映射
    model_to_type: dict[str, str] = {}
    if arrival_path and os.path.isfile(arrival_path):
        try:
            wb_a = openpyxl.load_workbook(arrival_path, read_only=True, data_only=True)
            ws_a = wb_a.active
            a_rows = list(ws_a.iter_rows(values_only=True))
            wb_a.close()
            for row in a_rows[1:]:
                if len(row) > 2 and row[1] and row[2]:
                    model_to_type[as_str(row[2])] = as_str(row[1])
        except Exception:
            pass

    # 设备位置表
    try:
        wb_p = openpyxl.load_workbook(position_path, read_only=True, data_only=True)
        sheet_name = "设备位置信息" if "设备位置信息" in wb_p.sheetnames else wb_p.sheetnames[-1]
        ws_p = wb_p[sheet_name]
        pos_rows = list(ws_p.iter_rows(values_only=True))
        wb_p.close()
    except Exception:
        return []

    groups: dict[tuple[str, str], list[dict]] = defaultdict(list)
    for row in pos_rows[1:]:
        if not row or not row[0]:
            continue
        model = as_str(row[0])
        if quotas:
            matched_spec = None
            for spec, remaining in quotas.items():
                if remaining <= 0:
                    continue
                if _model_allowed(model, {spec}):
                    matched_spec = spec
                    break
            if matched_spec is None:
                continue
            quotas[matched_spec] -= 1

        dev_name  = as_str(row[2]) if len(row) > 2 else ""
        room      = as_str(row[3]) if len(row) > 3 else ""
        col       = as_str(row[4]) if len(row) > 4 else ""
        u_pos     = as_str(row[5]) if len(row) > 5 else ""
        u_height  = as_str(row[6]) if len(row) > 6 else ""
        dev_class = get_sn_device_class(model, model_to_type)
        groups[(room, dev_class)].append({
            "设备大类":    dev_class,
            "厂家":       get_sn_vendor(model),
            "设备型号":   model,
            "设备名称":   dev_name,
            "所属机房":   room,
            "所属机柜":   col,
            "安装起始U位": u_pos,
            "设备U高":    u_height,
            "ESN":        "",
        })

    tables: list[dict] = []
    for (room, cls), rows in sorted(groups.items()):
        tables.append({
            "id":           f"{room}_{cls}",
            "room":         room,
            "device_class": cls,
            "file":         f"SN扫码表_{room}_{cls}.xlsx",
            "rows":         rows,
        })
    return tables


def generate_sn_xlsx(table: dict, output_dir: str) -> str:
    """为一个（机房, 设备大类）分组生成 SN 扫码表 xlsx，返回输出路径（失败返回空串）。"""
    try:
        import openpyxl  # type: ignore
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        return ""

    rows  = table.get("rows", [])
    fname = as_str(table.get("file")) or f"SN扫码表_{table.get('room', '')}_{table.get('device_class', '')}.xlsx"
    outpath = os.path.join(output_dir, fname)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SN扫码表"

    headers = ["序号", "设备大类", "厂家", "设备型号", "设备名称", "所属机房", "所属机柜", "安装起始U位", "设备U高", "ESN（待填）"]
    hf    = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(color="FFFFFF", bold=True, name="微软雅黑")
    esn_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hf; c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin

    for ri, row_data in enumerate(rows, 2):
        ws.cell(row=ri, column=1, value=ri - 1).border = thin
        for ci, key in enumerate(SN_COL_KEYS, 2):
            c = ws.cell(row=ri, column=ci, value=row_data.get(key, ""))
            c.border = thin
            c.alignment = Alignment(vertical="center")
            if key == "ESN":
                c.fill = esn_fill

    for ci, w in enumerate([6, 14, 10, 22, 40, 10, 10, 10, 8, 28], 1):
        ws.column_dimensions[chr(ord("A") + ci - 1)].width = w
    ws.row_dimensions[1].height = 22

    os.makedirs(os.path.dirname(outpath), exist_ok=True)
    wb.save(outpath)
    return outpath


def read_filled_sn_table(xlsx_path: str) -> list[dict]:
    """回读用户填好的 SN 扫码表，返回行 dict 列表（含 ESN）。"""
    try:
        import openpyxl  # type: ignore
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception:
        return []
    if not all_rows:
        return []

    header = [as_str(v) for v in all_rows[0]]
    # ESN 列名可能是 "ESN" 或 "ESN（待填）"
    esn_i = None
    for i, h in enumerate(header):
        if h.startswith("ESN") or h.upper() == "ESN":
            esn_i = i
            break

    def _at(row: tuple, name: str) -> str:
        for i, h in enumerate(header):
            if h == name and i < len(row):
                return as_str(row[i])
        return ""

    out: list[dict] = []
    for row in all_rows[1:]:
        if not row or all(v is None for v in row):
            continue
        esn = as_str(row[esn_i]) if (esn_i is not None and esn_i < len(row)) else ""
        out.append({
            "设备大类":   _at(row, "设备大类"),
            "厂家":      _at(row, "厂家"),
            "设备型号":  _at(row, "设备型号"),
            "设备名称":  _at(row, "设备名称"),
            "所属机房":  _at(row, "所属机房"),
            "所属机柜":  _at(row, "所属机柜"),
            "安装起始U位": _at(row, "安装起始U位"),
            "设备U高":   _at(row, "设备U高"),
            "ESN":      esn,
        })
    return out


def validate_esn(tables: list[dict]) -> tuple[list[str], dict[str, list[str]]]:
    """校验 ESN 完整性 + 唯一性。返回 (missing_items, dup_esns)。"""
    missing: list[str] = []
    dup: dict[str, list[str]] = {}
    seen: dict[str, str] = {}
    for tbl in tables:
        label = f"{as_str(tbl.get('room'))} {as_str(tbl.get('device_class'))}"
        for row in tbl.get("rows", []):
            esn = as_str(row.get("ESN"))
            if not esn:
                missing.append(
                    f"[{label}] 型号 {row.get('设备型号','')} / 机柜 {row.get('所属机柜','')} / U位 {row.get('安装起始U位','')}"
                )
            elif esn in seen:
                dup.setdefault(esn, [seen[esn]]).append(label)
            else:
                seen[esn] = label
    return missing, dup
