"""
plan_parser · 解析《任务计划表.xlsx》与《责任人信息表.xlsx》。

- parse_task_plan：提取活动 ID 以 "7." 开头的三级安装任务（不含 7.0 汇总行）。
- parse_principal_table：解析责任人信息表，返回 {unit::activity_id / unit::activity_name → {principal, principal_org}}。
"""
from __future__ import annotations

import hashlib
from typing import Any

from ._common import as_str, col_idx, parse_date_str


def is_rollup_activity_id(activity_id: str) -> bool:
    """活动 ID 7.0 为 7.x 的汇总行，不纳入三级安装任务。"""
    aid = as_str(activity_id).replace(" ", "")
    if not aid.startswith("7."):
        return False
    return aid.rstrip("0").rstrip(".") == "7"


def parse_task_plan(xlsx_path: str) -> list[dict]:
    """解析《任务计划表.xlsx》，提取活动 ID 以 "7." 开头的安装任务行（不含 7.0 汇总行）。"""
    try:
        import openpyxl  # type: ignore
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception:
        return []

    if not all_rows:
        return []

    header = [as_str(v) for v in all_rows[0]]
    ci_sn     = col_idx(header, "序列号")
    ci_aid    = col_idx(header, "活动ID", "级ID", "活动 ID")
    ci_unit   = col_idx(header, "管理单元")
    ci_name   = col_idx(header, "活动名称")
    ci_start  = col_idx(header, "开始日期")
    ci_end    = col_idx(header, "结束日期")
    ci_sla    = col_idx(header, "SLA")
    ci_dep    = col_idx(header, "依赖活动")
    ci_batch  = col_idx(header, "批次")
    ci_dev    = col_idx(header, "设备型号&数量的列表", "设备型号&数量列表", "设备型号")
    ci_remote = col_idx(header, "远程团队")
    ci_local  = col_idx(header, "现场团队", "本地团队")
    ci_owner  = col_idx(header, "责任人")

    if ci_aid is None:
        return []

    def _v(row: tuple, idx: int | None) -> Any:
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    tasks: list[dict] = []
    for row in all_rows[1:]:
        aid = as_str(_v(row, ci_aid))
        if not aid.startswith("7."):
            continue
        if is_rollup_activity_id(aid):
            continue
        sn    = as_str(_v(row, ci_sn))
        unit  = as_str(_v(row, ci_unit)) or "默认"
        name  = as_str(_v(row, ci_name)) or aid
        start = parse_date_str(_v(row, ci_start))
        end   = parse_date_str(_v(row, ci_end))
        tasks.append({
            "id":            sn or hashlib.sha256(f"{aid}:{unit}:{name}".encode()).hexdigest()[:10],
            "activity_id":   aid,
            "unit":          unit,
            "activity_name": name,
            "start_date":    start,
            "end_date":      end,
            "sla":           as_str(_v(row, ci_sla)),
            "dependencies":  as_str(_v(row, ci_dep)),
            "batch":         as_str(_v(row, ci_batch)),
            "devices":       as_str(_v(row, ci_dev)),
            "remote_team":   as_str(_v(row, ci_remote)),
            "local_team":    as_str(_v(row, ci_local)),
            "owner":         as_str(_v(row, ci_owner)),  # 原始责任人（可由责任人信息表覆盖）
            "principal":     "",        # 由责任人信息表填充
            "principal_org": "",        # 由责任人信息表填充
            "status":        "待下发",
            "progress_records": [],
        })

    return tasks


def parse_principal_table(xlsx_path: str) -> dict[str, dict]:
    """解析《责任人信息表.xlsx》。
    返回 {"{unit}::{activity_id}" / "{unit}::{activity_name}" : {principal, principal_org}}。
    """
    try:
        import openpyxl  # type: ignore
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception:
        return {}

    if not rows:
        return {}

    header = [as_str(v) for v in rows[0]]
    ci_unit  = col_idx(header, "管理单元")
    ci_aid   = col_idx(header, "活动ID")
    ci_name  = col_idx(header, "活动名称")
    ci_owner = col_idx(header, "责任人姓名", "责任人")
    ci_org   = col_idx(header, "责任主体（华为/分包商）", "责任主体")

    result: dict[str, dict] = {}
    for row in rows[1:]:
        def _v(idx: int | None) -> str:
            if idx is None or idx >= len(row):
                return ""
            return as_str(row[idx])

        unit  = _v(ci_unit)
        info  = {"principal": _v(ci_owner), "principal_org": _v(ci_org)}
        aid   = _v(ci_aid)
        aname = _v(ci_name)
        if aid:
            result[f"{unit}::{aid}"] = info
        if aname:
            result[f"{unit}::{aname}"] = info

    return result


def merge_principals(tasks: list[dict], principals: dict[str, dict]) -> int:
    """把责任人信息并入任务列表，返回更新条数。"""
    updated = 0
    for t in tasks:
        unit = as_str(t.get("unit"))
        info = (
            principals.get(f"{unit}::{as_str(t.get('activity_id'))}")
            or principals.get(f"{unit}::{as_str(t.get('activity_name'))}")
        )
        if info and (info.get("principal") or info.get("principal_org")):
            t["principal"] = info.get("principal", "") or t.get("principal", "")
            t["principal_org"] = info.get("principal_org", "") or t.get("principal_org", "")
            updated += 1
    return updated
