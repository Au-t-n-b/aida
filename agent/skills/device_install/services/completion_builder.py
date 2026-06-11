"""
completion_builder · 完工清单 & 完工报告生成。

- generate_completion_checklist：按（机房, 设备大类）各一份完工清单（含 ESN + 签字区）。
- generate_completion_report：全项目汇总（厂家/设备大类/型号/数量 + 签字区）。
"""
from __future__ import annotations

import os

from ._common import as_str

_HEADERS = ["序号", "设备大类", "厂家", "设备型号", "设备名称", "所属机房", "所属机柜", "安装起始U位", "设备U高", "ESN"]
_COL_KEYS = ["设备大类", "厂家", "设备型号", "设备名称", "所属机房", "所属机柜", "安装起始U位", "设备U高", "ESN"]


def generate_completion_checklist(table: dict, output_dir: str) -> str:
    """按（机房, 设备大类）生成《完工清单》xlsx，返回输出绝对路径（失败返回空串）。"""
    try:
        import openpyxl  # type: ignore
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        return ""

    rows       = table.get("rows", [])
    room       = as_str(table.get("room"))
    device_cls = as_str(table.get("device_class"))
    outpath    = os.path.join(output_dir, f"完工清单_{room}_{device_cls}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "完工清单"

    hf    = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(color="FFFFFF", bold=True, name="微软雅黑")
    thin  = Border(left=Side(style="thin"), right=Side(style="thin"),
                   top=Side(style="thin"), bottom=Side(style="thin"))
    alt   = PatternFill("solid", fgColor="EBF3FB")

    for ci, h in enumerate(_HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hf; c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin

    for ri, row_data in enumerate(rows, 2):
        ws.cell(row=ri, column=1, value=ri - 1).border = thin
        for ci, key in enumerate(_COL_KEYS, 2):
            c = ws.cell(row=ri, column=ci, value=row_data.get(key, ""))
            c.border = thin
            c.alignment = Alignment(vertical="center")
            if ri % 2 == 0:
                c.fill = alt

    sig_start  = len(rows) + 7
    for i, label in enumerate(["安装质量是否通过", "客户签字", "客户单位", "时间", "厂商签字", "时间"]):
        ws.cell(row=sig_start + i, column=len(_HEADERS), value=label).font = Font(name="微软雅黑")

    for ci, w in enumerate([6, 14, 10, 22, 40, 10, 10, 10, 8, 28], 1):
        ws.column_dimensions[chr(ord("A") + ci - 1)].width = w
    ws.row_dimensions[1].height = 22

    os.makedirs(os.path.dirname(outpath), exist_ok=True)
    wb.save(outpath)
    return outpath


def generate_completion_report(tables: list[dict], output_dir: str) -> str:
    """汇总所有完工清单设备行，按（厂家, 设备大类, 设备型号）计数生成《设备安装完工报告》。"""
    try:
        import openpyxl  # type: ignore
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from collections import Counter
    except ImportError:
        return ""

    counter: Counter = Counter()
    for table in tables:
        for row in table.get("rows", []):
            key = (as_str(row.get("厂家")), as_str(row.get("设备大类")), as_str(row.get("设备型号")))
            if key[1] or key[2]:
                counter[key] += 1

    sorted_items = sorted(counter.items(), key=lambda x: (x[0][1], x[0][2], x[0][0]))
    outpath = os.path.join(output_dir, "设备安装完工报告.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "设备安装完工报告"

    headers = ["序号", "厂家", "设备大类", "设备型号", "数量", ""]
    hf    = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(color="FFFFFF", bold=True, name="微软雅黑")
    thin  = Border(left=Side(style="thin"), right=Side(style="thin"),
                   top=Side(style="thin"), bottom=Side(style="thin"))
    alt   = PatternFill("solid", fgColor="EBF3FB")

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hf; c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin

    for ri, ((vendor, cls, model), qty) in enumerate(sorted_items, 2):
        for ci, v in enumerate([ri - 1, vendor, cls, model, qty, ""], 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = thin
            c.alignment = Alignment(vertical="center")
            if ri % 2 == 0:
                c.fill = alt

    sig_start  = len(sorted_items) + 7
    sig_labels = ["安装质量是否通过", "", "", "客户签字", "", "", "客户单位", "", "", "时间", "", "", "厂商签字", "", "", "时间"]
    for i, label in enumerate(sig_labels):
        if label:
            ws.cell(row=sig_start + i, column=6, value=label).font = Font(name="微软雅黑")

    for ci, w in enumerate([6, 14, 18, 28, 8, 20], 1):
        ws.column_dimensions[chr(ord("A") + ci - 1)].width = w
    ws.row_dimensions[1].height = 22

    os.makedirs(os.path.dirname(outpath), exist_ok=True)
    wb.save(outpath)
    return outpath
