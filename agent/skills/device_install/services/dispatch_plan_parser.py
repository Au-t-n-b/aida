"""
dispatch_plan_parser · 解析上游交付的自包含《设备安装实施计划.xlsx》。

Sheet「设备安装实施计划」→ tasks（写入 tasks_state.json）
Sheet「SN扫码表」         → 扁平行（写入 sn_pool.json，sn_generate 按勾选管理单元过滤）
"""
from __future__ import annotations

import hashlib
import json
from pathlib import Path

from ._common import as_str, col_idx

DISPATCH_PLAN_FILENAME = "设备安装实施计划.xlsx"
PLAN_SHEET = "设备安装实施计划"
SN_SHEET = "SN扫码表"


def resolve_dispatch_plan_path(source_dir: Path | str) -> Path | None:
    """在源目录查找《设备安装实施计划.xlsx》。"""
    p = Path(source_dir) / DISPATCH_PLAN_FILENAME
    return p if p.is_file() else None


def parse_dispatch_plan_with_sn(xlsx_path: str | Path) -> tuple[list[dict], list[dict]]:
    """解析合并实施计划 → (tasks, sn_rows)。

    tasks 字段对齐 tasks_state 任务 schema；status 缺省为「待下发」。
    sn_rows 保留 Sheet2 全部列（含 所属管理单元），ESN 列名兼容「ESN（待填）」。
    """
    try:
        import openpyxl  # type: ignore
    except ImportError as e:
        raise RuntimeError("解析实施计划需要 openpyxl") from e

    path = Path(xlsx_path)
    if not path.is_file():
        raise RuntimeError(f"实施计划文件不存在：{path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if PLAN_SHEET in wb.sheetnames:
            ws_plan = wb[PLAN_SHEET]
        else:
            ws_plan = wb.active
        plan_rows = list(ws_plan.iter_rows(values_only=True))

        ws_sn = wb[SN_SHEET] if SN_SHEET in wb.sheetnames else None
        sn_rows_raw = list(ws_sn.iter_rows(values_only=True)) if ws_sn else []
    finally:
        wb.close()

    tasks = _parse_plan_sheet(plan_rows)
    sn_rows = _parse_sn_sheet(sn_rows_raw)
    return tasks, sn_rows


def _parse_plan_sheet(all_rows: list[tuple]) -> list[dict]:
    if not all_rows:
        raise RuntimeError("实施计划 Sheet 为空")

    header = [as_str(v) for v in all_rows[0]]
    ci_plan_id = col_idx(header, "计划行ID")
    ci_unit    = col_idx(header, "管理单元")
    ci_aid     = col_idx(header, "活动ID", "活动 ID")
    ci_name    = col_idx(header, "活动名称")
    ci_start   = col_idx(header, "计划开始", "开始日期")
    ci_end     = col_idx(header, "计划完成", "计划结束", "结束日期")
    ci_sla     = col_idx(header, "SLA")
    ci_pri     = col_idx(header, "责任人", "责任人姓名")
    ci_org     = col_idx(header, "责任主体", "责任主体（华为/分包商）")
    ci_dev     = col_idx(header, "设备型号&数量", "设备型号&数量的列表")
    ci_status  = col_idx(header, "状态")

    if ci_unit is None and ci_aid is None:
        raise RuntimeError("实施计划 Sheet 缺少「管理单元」或「活动ID」列")

    def _v(row: tuple, idx: int | None) -> str:
        if idx is None or idx >= len(row):
            return ""
        return as_str(row[idx])

    tasks: list[dict] = []
    for row in all_rows[1:]:
        if not row or all(v is None or as_str(v) == "" for v in row):
            continue
        unit = _v(row, ci_unit)
        aid  = _v(row, ci_aid)
        name = _v(row, ci_name)
        if not unit and not aid and not name:
            continue
        plan_id = _v(row, ci_plan_id) or f"{unit}::{aid}"
        tid = plan_id or hashlib.sha256(f"{aid}:{unit}:{name}".encode()).hexdigest()[:10]
        tasks.append({
            "id":            tid,
            "plan_row_id":   plan_id,
            "activity_id":   aid,
            "unit":          unit or "默认",
            "activity_name": name or aid,
            "start_date":    _v(row, ci_start),
            "end_date":      _v(row, ci_end),
            "sla":           _v(row, ci_sla),
            "principal":     _v(row, ci_pri),
            "principal_org": _v(row, ci_org),
            "devices":       _v(row, ci_dev),
            "dependencies":  "",
            "batch":         "",
            "remote_team":   "",
            "local_team":    "",
            "owner":         _v(row, ci_pri),
            "status":        _v(row, ci_status) or "待下发",
            "progress_records": [],
        })

    if not tasks:
        raise RuntimeError("实施计划 Sheet 未解析到任何任务行")
    return tasks


def _parse_sn_sheet(all_rows: list[tuple]) -> list[dict]:
    if not all_rows:
        return []

    header = [as_str(v) for v in all_rows[0]]
    keys = [
        ("所属管理单元", ("所属管理单元",)),
        ("关联计划行ID", ("关联计划行ID", "关联活动", "关联计划")),
        ("设备大类", ("设备大类",)),
        ("厂家", ("厂家",)),
        ("设备型号", ("设备型号",)),
        ("设备名称", ("设备名称",)),
        ("所属机房", ("所属机房",)),
        ("所属机柜", ("所属机柜",)),
        ("安装起始U位", ("安装起始U位",)),
        ("设备U高", ("设备U高",)),
        ("ESN", ("ESN", "ESN（待填）")),
    ]
    col_map: dict[str, int | None] = {k: col_idx(header, *aliases) for k, aliases in keys}

    rows_out: list[dict] = []
    for row in all_rows[1:]:
        if not row or all(v is None or as_str(v) == "" for v in row):
            continue
        rec: dict[str, str] = {}
        for k, idx in col_map.items():
            if idx is not None and idx < len(row):
                rec[k] = as_str(row[idx])
            else:
                rec[k] = ""
        if not any(rec.values()):
            continue
        rows_out.append(rec)
    return rows_out


def group_sn_rows_to_tables(sn_rows: list[dict]) -> list[dict]:
    """将扁平行按 (所属机房, 设备大类) 分组，结构对齐 sn_tables.json 的 tables[]。"""
    from collections import defaultdict

    groups: dict[tuple[str, str], list[dict]] = defaultdict(list)
    for r in sn_rows:
        room = as_str(r.get("所属机房"))
        cls  = as_str(r.get("设备大类")) or "其他"
        groups[(room, cls)].append({
            "设备大类":    cls,
            "厂家":       as_str(r.get("厂家")),
            "设备型号":   as_str(r.get("设备型号")),
            "设备名称":   as_str(r.get("设备名称")),
            "所属机房":   room,
            "所属机柜":   as_str(r.get("所属机柜")),
            "安装起始U位": as_str(r.get("安装起始U位")),
            "设备U高":    as_str(r.get("设备U高")),
            "ESN":        as_str(r.get("ESN")),
            "所属管理单元": as_str(r.get("所属管理单元")),
            "关联计划行ID": as_str(r.get("关联计划行ID")),
        })

    tables: list[dict] = []
    for (room, cls), rows in sorted(groups.items()):
        tables.append({
            "id":           f"{room}_{cls}",
            "room":         room,
            "device_class": cls,
            "file":         f"SN扫码表_{room}_{cls}.xlsx",
            "rows":         rows,
        })
    return tables


def save_sn_pool(path: str | Path, *, source: str, rows: list[dict]) -> None:
    """写入 sn_pool.json（全量 SN，供 sn_generate 按下发管理单元过滤）。"""
    from .task_store import iso_now

    payload = {
        "loaded_at": iso_now(),
        "source": str(source),
        "rows": rows,
    }
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def load_sn_pool(path: str | Path) -> list[dict]:
    p = Path(path)
    if not p.is_file():
        return []
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return []
    rows = data.get("rows") or []
    return [r for r in rows if isinstance(r, dict)]
