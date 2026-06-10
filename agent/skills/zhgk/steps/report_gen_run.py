"""
report_gen_run · 工勘报告生成（9 表 Word）

意图: report_gen 专属

流程:
  1. 读取全量勘测结果表（需已含 AI 评估结果）
  2. 重建 AssessmentResult 列表（从 AI评估结果 列）
  3. load_risk_library + identify_risks + write_risk_table
  4. 从 Output/ 读取问题清单表（如存在）
  5. 构建 ProjectMeta（from project dict + project_info.json）
  6. build_report() → 工勘报告.docx
  7. 返回 artifacts + metrics
"""
from __future__ import annotations

import json
import os

from ...base import BaseStep, SkillContext, SkillState, StepResult, Emit, CheckResult
from ._intent_guard import should_skip


def _get_survey_table(ctx: SkillContext) -> str | None:
    info_path = ctx.runtime_dir / "project_info.json"
    if info_path.exists():
        try:
            path = json.loads(info_path.read_text(encoding="utf-8")).get("survey_table_path", "")
            if path and os.path.exists(path):
                return path
        except Exception:
            pass
    tables = sorted(ctx.output_dir.glob("*全量勘测结果表*.xlsx")) if ctx.output_dir.exists() else []
    return str(tables[0]) if tables else None


def _get_generation_cooling(ctx: SkillContext) -> str:
    gc = ctx.project.get("generation_cooling", "")
    if gc:
        return gc
    info_path = ctx.runtime_dir / "project_info.json"
    if info_path.exists():
        try:
            return json.loads(info_path.read_text(encoding="utf-8")).get("generation_cooling", "")
        except Exception:
            pass
    return ""


def _read_issue_list_file(output_dir) -> list[dict]:
    """从 Output/ 读取问题清单表，返回 IssueItem 列表"""
    import openpyxl
    files = sorted(output_dir.glob("*问题清单表*.xlsx")) if output_dir.exists() else []
    if not files:
        return []
    try:
        wb = openpyxl.load_workbook(str(files[0]), read_only=True, data_only=True)
        ws = wb.active
        headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
        col_map = {name: idx for idx, name in enumerate(headers) if name}
        items = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            def _s(key: str) -> str:
                idx = col_map.get(key)
                return str(row[idx] or "").strip() if idx is not None else ""
            items.append({
                "序号": int(row[0]) if row[0] else len(items) + 1,
                "问题描述": _s("问题描述"),
                "状态": _s("状态") or "open",
                "整改建议": _s("整改建议"),
                "责任人": "",
                "计划关闭时间": "",
                "备注": _s("备注"),
            })
        wb.close()
        return items
    except Exception:
        return []


def _rebuild_assessment_results(survey_rows: list[dict]) -> list:
    """从勘测结果行重建 AssessmentResult 列表"""
    from ..services.types import AssessmentResult, AssessmentValue
    valid_values = {v.value for v in AssessmentValue}
    results = []
    for row in survey_rows:
        ai_val = (row.get("AI评估结果") or "").strip()
        if ai_val in valid_values:
            conclusion = AssessmentValue.from_str(ai_val)
        else:
            conclusion = AssessmentValue.NOT_SURVEYED
        results.append(AssessmentResult(conclusion=conclusion, defect_description="", confidence=0.9))
    return results


class ReportGenRunStep(BaseStep):
    key = "report_gen_run"
    name = "报告生成"
    artifacts_pattern = [
        "ProjectData/Output/工勘报告.docx",
        "ProjectData/Output/*风险识别结果表*.xlsx",
    ]

    def check_inputs(self, ctx: SkillContext) -> CheckResult:
        if should_skip(self.key, ctx.project):
            return {"ok": True, "missing": []}

        missing = []
        if _get_survey_table(ctx) is None:
            missing.append("ProjectData/Output/*全量勘测结果表*.xlsx")

        from ..path_config import get_report_template_path, get_risk_library_path
        if not os.path.exists(get_report_template_path()):
            missing.append("ProjectData/Template/新版项目工勘报告模板.docx")
        if not os.path.exists(get_risk_library_path()):
            missing.append("ProjectData/Template/工勘常见高风险库.xlsx")

        return {"ok": not missing, "missing": missing}

    def run(self, ctx: SkillContext, state: SkillState, emit: Emit) -> StepResult:
        if should_skip(self.key, ctx.project):
            return {}

        from ..services.survey_table_builder import read_survey_table
        from ..services.risk_engine import load_risk_library, identify_risks, write_risk_table
        from ..services.report_builder import build_report
        from ..services.types import ProjectMeta
        from ..services._llm_adapter import make_llm_adapter
        from ..path_config import get_report_template_path, get_risk_library_path

        survey_table = _get_survey_table(ctx)
        if not survey_table:
            raise RuntimeError("report_gen_run: 全量勘测结果表不存在")

        proj = ctx.project
        gen_cooling = _get_generation_cooling(ctx)
        output_dir = str(ctx.output_dir)
        llm = make_llm_adapter(ctx, step_key=self.key)

        # ── 1. 读取勘测结果 ─────────────────────────────────────────────
        emit(f"[report_gen_run] 读取勘测结果表: {os.path.basename(survey_table)}")
        survey_rows = read_survey_table(survey_table)
        emit(f"[report_gen_run] 总条目: {len(survey_rows)}")

        assessment_results = _rebuild_assessment_results(survey_rows)

        # ── 2. 风险识别 ──────────────────────────────────────────────────
        risk_library_path = get_risk_library_path()
        emit(f"[report_gen_run] 加载高风险库: {os.path.basename(risk_library_path)}")
        risk_items_raw = load_risk_library(risk_library_path, generation_cooling=gen_cooling)
        emit(f"[report_gen_run] 高风险条目: {len(risk_items_raw)} 条")

        emit("[report_gen_run] LLM 风险判断中…")
        triggered_risks = identify_risks(risk_items_raw, survey_rows, llm)
        emit(f"[report_gen_run] 触发风险: {len(triggered_risks)} 条")

        risk_table_path = write_risk_table(
            triggered_risks,
            output_dir=output_dir,
            activity_id=proj.get("activity_id", ""),
            project_name=proj.get("project_name", ""),
            room_name=proj.get("room_name", ""),
        )
        emit(f"[report_gen_run] 风险表: {os.path.basename(risk_table_path)}")

        # ── 3. 读取问题清单 ──────────────────────────────────────────────
        issue_list = _read_issue_list_file(ctx.output_dir)
        emit(f"[report_gen_run] 问题清单: {len(issue_list)} 条")

        # ── 4. 构建 ProjectMeta ──────────────────────────────────────────
        # 从 project_info.json 补充 survey_date / surveyor
        info: dict = {}
        info_path = ctx.runtime_dir / "project_info.json"
        if info_path.exists():
            try:
                info = json.loads(info_path.read_text(encoding="utf-8"))
            except Exception:
                pass

        project_meta = ProjectMeta(
            project_name=proj.get("project_name", info.get("project_name", "")),
            activity_id=proj.get("activity_id", info.get("activity_id", "")),
            room_name=proj.get("room_name", info.get("room_name", "")),
            survey_date=proj.get("survey_date", info.get("survey_date", "")),
            surveyor=proj.get("surveyor", info.get("surveyor", "")),
            generation_cooling=gen_cooling,
        )

        # ── 5. 生成报告 ──────────────────────────────────────────────────
        template_path = get_report_template_path()
        emit(f"[report_gen_run] 填充报告模板: {os.path.basename(template_path)}")

        report_path = build_report(
            template_path=template_path,
            output_dir=output_dir,
            project_meta=project_meta,
            survey_results=survey_rows,
            assessment_results=assessment_results,
            issue_list=issue_list,
            risk_list=triggered_risks,
            photo_map=None,
            llm_call=llm,
        )
        emit(f"[report_gen_run] ✓ 工勘报告: {os.path.basename(report_path)}")

        return {
            "metrics": {
                "total_items": len(survey_rows),
                "triggered_risks": len(triggered_risks),
                "issue_count": len(issue_list),
                "report_path": report_path,
            },
            "artifacts": [
                ctx.rel(report_path),
                ctx.rel(risk_table_path),
            ],
        }
