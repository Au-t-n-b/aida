"""
wait_survey · 等待现场勘测结果上传

意图: survey_work 专属

HITL FilePicker：等待勘测工程师填写全量勘测结果表并上传。
  - check_inputs: 检查主表是否已有勘测结果，或 Input/ 是否有新上传的填写表
  - run: 读取上传的填写表 → write_survey_results(round=N) → 删除上传文件

多轮复勘支持：
  - 若 project["resurvey_decision"] == "resurvey"，跳过「已有结果」检测，
    直接等待新上传（复勘第 N+1 轮）
  - resume 时 apply_resume_payload("wait_survey") 清除 resurvey_decision，
    确保下次进来走正常路径
"""
from __future__ import annotations

import json
import os

from ...base import BaseStep, SkillContext, SkillState, StepResult, Emit, CheckResult
from ._intent_guard import should_skip

# 用户需要将填好的表格放到此路径
UPLOADED_FILENAME = "已填写_全量勘测结果表.xlsx"


def _get_survey_table(ctx: SkillContext) -> str | None:
    info_path = ctx.runtime_dir / "project_info.json"
    if info_path.exists():
        try:
            path = json.loads(info_path.read_text(encoding="utf-8")).get("survey_table_path", "")
            if path and os.path.exists(path):
                return path
        except Exception:
            pass
    tables = sorted(ctx.output_dir.glob("*全量勘测结果表*.xlsx")) if ctx.output_dir.exists() else []
    return str(tables[0]) if tables else None


def _find_uploaded_table(ctx: SkillContext) -> str | None:
    """在 Input/ 查找用户上传的已填写表"""
    if ctx.input_dir.exists():
        # 精确名匹配
        exact = ctx.input_dir / UPLOADED_FILENAME
        if exact.exists():
            return str(exact)
        # 模糊匹配（含"全量勘测结果表"且不是 BOQ）
        for p in sorted(ctx.input_dir.glob("*全量勘测结果表*.xlsx")):
            if "boq" not in p.name.lower():
                return str(p)
    return None


def _has_survey_results(survey_table_path: str) -> bool:
    """检查主表中是否已有非空的「最新检查结果」（即已合并过某轮结果）"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(survey_table_path, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
        try:
            res_col = headers.index("最新检查结果") + 1
        except ValueError:
            wb.close()
            return False
        for row in ws.iter_rows(min_row=2, max_col=res_col, values_only=True):
            if row[-1] is not None and str(row[-1]).strip():
                wb.close()
                return True
        wb.close()
    except Exception:
        pass
    return False


class WaitSurveyStep(BaseStep):
    key = "wait_survey"
    name = "等待现场上传"
    artifacts_pattern = []

    def _gkclaw_status_note(self, ctx: SkillContext) -> str:
        """GKCLAW 链路钩子：有已下发任务时拉取邮件回传并返回状态行（异常不阻断流程）。

        dry-run 任务或 mailgw 未配置时只展示状态不拉取（边界 C9/C10：拉取发生在
        本 check 被调用时——run 启动与每次 resume；无后台轮询）。
        """
        try:
            info_path = ctx.runtime_dir / "project_info.json"
            if not info_path.exists():
                return ""
            info = json.loads(info_path.read_text(encoding="utf-8"))
            tid = info.get("gkclaw_task_id", "")
            if not tid:
                return ""
            from ..services.gkclaw.registry import TaskRegistry
            reg = TaskRegistry(ctx.runtime_dir)
            task = reg.get(tid)
            if task is None:
                return ""
            alerts: list[str] = []
            if not task.get("dry_run"):
                import agent.mailbox as mailbox
                if mailbox.is_configured() and task["state"] in (
                        "dispatched", "accepted", "staged_returned"):
                    from ..services.gkclaw.ingest import poll_and_ingest
                    summary = poll_and_ingest(
                        runtime_dir=ctx.runtime_dir, input_dir=ctx.input_dir,
                        survey_table_path=_get_survey_table(ctx))
                    alerts = list(summary.get("alerts", []))
                    task = reg.get(tid) or task
            bits = [f"GKCLAW 任务 {tid} · 状态 {task['state']}"]
            if task.get("dry_run"):
                bits.append("dry-run 未真发")
            if task.get("web_access_url"):
                bits.append(f"现场 Web 入口 {task['web_access_url']}")
            if task.get("merge_blocked"):
                bits.append(f"⚠ 合并阻塞：{task.get('merge_blocked_reason', '')}")
            return "；".join(bits) + ("；" + "；".join(alerts[-3:]) if alerts else "")
        except Exception as e:  # noqa: BLE001 — 邮件链路异常绝不阻断人工上传通道
            return f"[gkclaw] 拉取回传失败：{e}"

    def check_inputs(self, ctx: SkillContext) -> CheckResult:
        if should_skip(self.key, ctx.project):
            return {"ok": True, "missing": []}

        gk_note = self._gkclaw_status_note(ctx)
        resurvey_pending = ctx.project.get("resurvey_decision") == "resurvey"
        survey_table = _get_survey_table(ctx)

        if resurvey_pending:
            # 复勘模式：忽略「已有结果」检测，必须有新上传文件
            if _find_uploaded_table(ctx) is not None:
                return {"ok": True, "missing": []}
            base_note = (
                "请将第 N+1 轮复勘后的全量勘测结果表（填写「最新检查结果」列）"
                f"保存为 {UPLOADED_FILENAME} 并上传到 ProjectData/Input/"
            )
            return {
                "ok": False,
                "missing": [f"ProjectData/Input/{UPLOADED_FILENAME}"],
                "note": base_note + (f"\n{gk_note}" if gk_note else ""),
            }

        # 正常流程
        if survey_table and _has_survey_results(survey_table):
            return {"ok": True, "missing": []}

        if _find_uploaded_table(ctx) is not None:
            return {"ok": True, "missing": []}

        base_note = (
            "请从 Output/ 下载全量勘测结果表，完成现场勘测后填写「最新检查结果」列，"
            f"将文件保存为 {UPLOADED_FILENAME} 后上传到 ProjectData/Input/"
        )
        return {
            "ok": False,
            "missing": [f"ProjectData/Input/{UPLOADED_FILENAME}"],
            "note": base_note + (f"\n{gk_note}" if gk_note else ""),
        }

    def run(self, ctx: SkillContext, state: SkillState, emit: Emit) -> StepResult:
        if should_skip(self.key, ctx.project):
            return {}

        from ..services.survey_table_builder import read_survey_table
        from ..services.resurvey_manager import get_current_round, write_survey_results

        survey_table = _get_survey_table(ctx)
        if not survey_table:
            raise RuntimeError("wait_survey: 全量勘测结果表不存在")

        resurvey_pending = ctx.project.get("resurvey_decision") == "resurvey"

        # 非复勘且主表已有结果 → 无需重合并
        if not resurvey_pending and _has_survey_results(survey_table):
            emit("[wait_survey] ✓ 勘测结果已就绪（跳过合并）")
            return {"metrics": {"survey_already_filled": True}}

        uploaded = _find_uploaded_table(ctx)
        if not uploaded:
            raise RuntimeError(
                f"wait_survey: 未找到上传的勘测结果表（期望: Input/{UPLOADED_FILENAME}）"
            )

        emit(f"[wait_survey] 读取已填写表: {os.path.basename(uploaded)}")
        filled_rows = read_survey_table(uploaded)
        results: dict[int, str] = {
            row["序号"]: row["最新检查结果"]
            for row in filled_rows
            if row.get("最新检查结果", "").strip()
        }

        if not results:
            emit("[wait_survey] ⚠ 上传的表格「最新检查结果」列全部为空，请检查文件")
            return {"metrics": {"filled_count": 0}}

        round_num = get_current_round(survey_table)
        emit(f"[wait_survey] 合并第 {round_num} 轮勘测结果：{len(results)} 条")
        write_survey_results(survey_table, results, round_num)
        emit(f"[wait_survey] ✓ 结果已合并（第{round_num}轮，{len(results)} 条）")

        # 删除上传的临时文件，为下一轮复勘准备
        try:
            os.remove(uploaded)
            emit(f"[wait_survey] 已清理上传临时文件: {os.path.basename(uploaded)}")
        except Exception:
            pass

        # 多轮复勘历史：累积每轮汇总（SDUI 轮次对比 Table）。
        # metrics 扁平 merge 是 last-write-wins，故读旧历史→去重当前轮→追加→整列回写。
        history = list((state.get("metrics") or {}).get("survey_round_history") or [])
        history = [h for h in history if h.get("round") != round_num]
        history.append({
            "round": round_num,
            "filled": len(results),
            "total": len(filled_rows),
        })
        history.sort(key=lambda h: h.get("round", 0))

        return {
            "metrics": {
                "survey_round": round_num,
                "filled_count": len(results),
                "survey_round_history": history,
            }
        }
