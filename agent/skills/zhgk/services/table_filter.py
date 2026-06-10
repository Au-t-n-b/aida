"""
智慧工勘 v4 — 底表加载与过滤

职责: 加载入场评估标准表并按条件过滤
错误码前缀: SS-TF
"""
from __future__ import annotations

import os
from typing import Optional

import openpyxl

from .types import SurveyItem


# ──────────────────────────────────────────────
# 异常定义
# ──────────────────────────────────────────────

class TableFilterError(Exception):
    """底表过滤相关异常"""

    def __init__(self, code: str, message: str):
        self.code = code
        self.message = message
        super().__init__(f"[{code}] {message}")


# ──────────────────────────────────────────────
# 必要列名
# ──────────────────────────────────────────────

REQUIRED_COLUMNS = {"代际-制冷", "分类", "细分场景", "勘测要素", "检查内容"}

# 底表中的全部列名（用于按名称索引）
ALL_COLUMN_NAMES = [
    "序号", "代际-制冷", "分类", "细分场景", "勘测要素", "项目",
    "检查内容", "是否支持视频勘测", "勘测方法", "检查结果", "备注",
    "语音助手背景知识", "视频勘测背景知识",
]


# ──────────────────────────────────────────────
# 公开接口
# ──────────────────────────────────────────────

def load_base_table(
    path: str,
    sheet_name: str = "入场评估标准",
) -> list[SurveyItem]:
    """
    从入场评估标准表加载全量数据。

    跳过"代际-制冷"列为空的行（底表中用作空行分隔）。

    异常:
        TableFilterError(SS-TF-E-001): 文件不存在
        TableFilterError(SS-TF-E-002): Sheet 不存在
        TableFilterError(SS-TF-E-004): 必要列名缺失
    """
    if not os.path.exists(path):
        raise TableFilterError("SS-TF-E-001", f"底表文件不存在: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        wb.close()
        raise TableFilterError(
            "SS-TF-E-002", f"Sheet '{sheet_name}' 不存在，可用: {wb.sheetnames}"
        )

    ws = wb[sheet_name]

    # 读取表头（第1行）
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h).strip() if h else "" for h in header_row]

    # 验证必要列
    header_set = set(headers)
    missing = REQUIRED_COLUMNS - header_set
    if missing:
        wb.close()
        raise TableFilterError(
            "SS-TF-E-004", f"底表必要列缺失: {missing}"
        )

    # 建立列名→索引映射
    col_map = {name: idx for idx, name in enumerate(headers) if name}

    items: list[SurveyItem] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        gc_val = _cell_str(row, col_map.get("代际-制冷"))
        if not gc_val:
            continue  # 跳过空行分隔

        item: SurveyItem = {
            "序号": _cell_int(row, col_map.get("序号")),
            "代际_制冷": gc_val,
            "分类": _cell_str(row, col_map.get("分类")),
            "细分场景": _cell_str(row, col_map.get("细分场景")),
            "勘测要素": _cell_str(row, col_map.get("勘测要素")),
            "项目": _cell_str(row, col_map.get("项目")),
            "检查内容": _cell_str(row, col_map.get("检查内容")),
            "是否支持视频勘测": _cell_str(row, col_map.get("是否支持视频勘测")),
            "勘测方法": _cell_str(row, col_map.get("勘测方法")),
            "检查结果": _cell_str(row, col_map.get("检查结果")),
            "备注": _cell_str(row, col_map.get("备注")),
            "语音助手背景知识": _cell_str(row, col_map.get("语音助手背景知识")),
            "视频勘测背景知识": _cell_str(row, col_map.get("视频勘测背景知识")),
        }
        items.append(item)

    wb.close()
    return items


def filter_items(
    items: list[SurveyItem],
    generation_cooling: str,
    category: str = "标准",
    sub_scenes: Optional[list[str]] = None,
    data_keywords: Optional[list[str]] = None,
) -> list[SurveyItem]:
    """
    按条件过滤底表条目。

    过滤逻辑（全部 AND）:
        1. 代际_制冷 == generation_cooling
        2. 分类 == category
        3. (sub_scenes is None) OR (细分场景 in sub_scenes)
        4. (data_keywords is None) OR any(kw in 勘测要素 for kw in data_keywords)

    异常:
        TableFilterError(SS-TF-E-003): 过滤结果为空
    """
    result = []
    for item in items:
        if item["代际_制冷"] != generation_cooling:
            continue
        if item["分类"] != category:
            continue
        if sub_scenes is not None and item["细分场景"] not in sub_scenes:
            continue
        if data_keywords is not None:
            if not any(kw in item["勘测要素"] for kw in data_keywords):
                continue
        result.append(item)

    if not result:
        raise TableFilterError(
            "SS-TF-E-003",
            f"过滤结果为空: generation_cooling={generation_cooling}, "
            f"category={category}, sub_scenes={sub_scenes}, "
            f"data_keywords={data_keywords}",
        )

    return result


def get_available_data_keywords(
    items: list[SurveyItem],
    generation_cooling: str,
) -> list[str]:
    """
    获取指定代际-制冷下，分类=数据的所有勘测要素去重列表。
    供用户选择"追加哪些数据类条目"。
    """
    keywords = set()
    for item in items:
        if item["代际_制冷"] == generation_cooling and item["分类"] == "数据":
            if item["勘测要素"]:
                keywords.add(item["勘测要素"])
    return sorted(keywords)


def get_sub_scenes_for_cooling(generation_cooling: str) -> list[str]:
    """
    根据代际-制冷返回默认的细分场景列表。

    规则:
        - 含"风冷": ["硬装入场"]
        - 含"液冷": ["硬装入场", "通液前", "加电前"]
    """
    if "液冷" in generation_cooling:
        return ["硬装入场", "通液前", "加电前"]
    return ["硬装入场"]


# ──────────────────────────────────────────────
# 内部工具
# ──────────────────────────────────────────────

def _cell_str(row: tuple, idx: Optional[int]) -> str:
    """安全获取单元格字符串值。"""
    if idx is None or idx >= len(row):
        return ""
    val = row[idx]
    if val is None:
        return ""
    return str(val).strip()


def _cell_int(row: tuple, idx: Optional[int]) -> int:
    """安全获取单元格整数值。"""
    if idx is None or idx >= len(row):
        return 0
    val = row[idx]
    if val is None:
        return 0
    try:
        return int(val)
    except (ValueError, TypeError):
        return 0
