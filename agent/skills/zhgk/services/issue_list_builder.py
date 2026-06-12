"""
智慧工勘 v4 — 问题清单构建

职责: 从全量勘测结果表生成问题清单表
错误码前缀: SS-IL
"""
from __future__ import annotations

import json
import os

import openpyxl

from .logger import log_error, log_info, log_warn
from .types import IssueGenResult, IssueStatus, LLMCallable


# ──────────────────────────────────────────────
# 异常定义
# ──────────────────────────────────────────────

class IssueListError(Exception):
    def __init__(self, code: str, message: str):
        self.code = code
        self.message = message
        super().__init__(f"[{code}] {message}")


# ──────────────────────────────────────────────
# Prompt 模板
# ──────────────────────────────────────────────

ISSUE_SYSTEM_PROMPT = """你是一名数据中心工勘问题分析专家。根据勘测检查项的「检查内容」和「检查结果」，
生成该问题项的描述、整改建议，并判定严重度。

输出格式（严格 JSON）：
{
  "problem_description": "简洁的问题描述（1-2句话，说明什么不满足）",
  "remediation_suggestion": "具体可操作的整改建议（1-3句话）",
  "severity": "high | mid | low"
}

严重度判定标准：
- high（高）：涉及安全、合规、承重、供电/制冷等关键项严重不达标，直接阻断交付或存在安全隐患
- mid（中）：明确不满足标准、但可通过整改解决，不阻断交付
- low（低）：轻微偏差，或仅因数据缺失/图片不清「无法识别」需补充确认

注意：
1. 问题描述应具体，指出实际值与标准值的差距
2. 整改建议应可操作，包含具体动作而非泛泛而谈
3. severity 必须是 high / mid / low 之一
4. 只输出 JSON"""

ISSUE_USER_TEMPLATE = (
    "检查内容: {check_content}\n"
    "检查结果: {latest_result}\n"
    "评估结论: {assessment_conclusion}"
)


# ──────────────────────────────────────────────
# 问题清单表头
# ──────────────────────────────────────────────

ISSUE_TABLE_HEADERS = [
    "序号", "严重度", "问题描述", "状态", "整改建议", "责任人", "计划关闭时间", "备注",
]

# 触发问题的评估结论
ISSUE_TRIGGER_VALUES = {"不满足", "无法识别"}

# 严重度枚举 → 中文（写表 / KPI 展示用；枚举值对齐前端 RiskList 的 high/mid/low）
_VALID_SEVERITY = ("high", "mid", "low")
_SEVERITY_CN = {"high": "高", "mid": "中", "low": "低"}


# ──────────────────────────────────────────────
# 公开接口
# ──────────────────────────────────────────────

def build_issue_list(
    survey_table_path: str,
    output_dir: str,
    activity_id: str,
    project_name: str,
    room_name: str,
    llm_call: LLMCallable,
) -> tuple[str, list[dict]]:
    """
    从全量勘测结果表生成问题清单表。

    筛选条件: AI评估结果 = "不满足" 或 "无法识别"

    返回:
        (生成的问题清单表文件路径, 问题摘要行列表)
        摘要行: [{"序号", "问题描述", "状态", "整改建议"}]，供 SDUI 投影问题清单 Table。
    """
    if not os.path.exists(survey_table_path):
        raise IssueListError("SS-IL-E-003", f"全量勘测结果表不存在: {survey_table_path}")

    # 读取全量勘测结果表
    wb_src = openpyxl.load_workbook(survey_table_path, read_only=True, data_only=True)
    ws_src = wb_src.active

    header_row = [str(ws_src.cell(1, c).value or "").strip() for c in range(1, ws_src.max_column + 1)]
    col_map = {name: idx for idx, name in enumerate(header_row) if name}

    if "AI评估结果" not in col_map:
        wb_src.close()
        raise IssueListError("SS-IL-E-001", "全量勘测结果表无 AI评估结果 列（需先执行评估）")

    check_idx = col_map.get("检查内容")
    result_idx = col_map.get("最新检查结果")
    ai_idx = col_map["AI评估结果"]

    # 筛选不满足/无法识别的行
    issue_rows = []
    for row in ws_src.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        ai_val = str(row[ai_idx] or "").strip()
        if ai_val in ISSUE_TRIGGER_VALUES:
            issue_rows.append({
                "check_content": str(row[check_idx] or "").strip() if check_idx is not None else "",
                "latest_result": str(row[result_idx] or "").strip() if result_idx is not None else "",
                "assessment": ai_val,
            })

    wb_src.close()

    # 对每行调用 LLM 生成问题描述
    issues: list[IssueGenResult] = []
    for item in issue_rows:
        try:
            result = _generate_issue(item, llm_call)
            issues.append(result)
        except Exception as e:
            log_warn("issue_list_builder", "build_issue_list", f"LLM 生成失败: {e}")
            issues.append(IssueGenResult(
                problem_description=f"[自动生成失败] {item['check_content']}",
                remediation_suggestion="请人工补充整改建议",
            ))

    # 创建问题清单表
    os.makedirs(output_dir, exist_ok=True)
    filename = f"{activity_id}_{project_name}_{room_name}_问题清单表.xlsx"
    output_path = os.path.join(output_dir, filename)

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "问题清单"

    for col_idx, header in enumerate(ISSUE_TABLE_HEADERS, 1):
        ws_out.cell(row=1, column=col_idx, value=header)

    summary_rows: list[dict] = []
    for i, issue in enumerate(issues, 1):
        sev_cn = _SEVERITY_CN.get(issue.severity, "中")
        ws_out.cell(row=i + 1, column=1, value=i)
        ws_out.cell(row=i + 1, column=2, value=sev_cn)
        ws_out.cell(row=i + 1, column=3, value=issue.problem_description)
        ws_out.cell(row=i + 1, column=4, value=IssueStatus.OPEN.value)
        ws_out.cell(row=i + 1, column=5, value=issue.remediation_suggestion)
        ws_out.cell(row=i + 1, column=6, value="")
        ws_out.cell(row=i + 1, column=7, value="")
        ws_out.cell(row=i + 1, column=8, value="")
        summary_rows.append({
            "序号": i,
            "严重度": sev_cn,          # 中文，供 SDUI 问题表首列色点
            "severity": issue.severity,  # 枚举 high/mid/low，供 step 统计分级
            "问题描述": issue.problem_description,
            "状态": IssueStatus.OPEN.value,
            "整改建议": issue.remediation_suggestion,
        })

    wb_out.save(output_path)
    wb_out.close()

    log_info("issue_list_builder", "build_issue_list",
             f"问题清单已生成: {len(issues)} 条问题 → {output_path}")
    return output_path, summary_rows


def update_issue_status(
    issue_list_path: str,
    row_index: int,
    new_status: IssueStatus,
) -> None:
    """更新问题清单中某行的状态。row_index 为 1-based 数据行号。"""
    if not os.path.exists(issue_list_path):
        raise IssueListError("SS-IL-E-003", f"问题清单文件不存在: {issue_list_path}")

    wb = openpyxl.load_workbook(issue_list_path)
    ws = wb.active

    header_row = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    status_col = None
    for idx, h in enumerate(header_row):
        if h == "状态":
            status_col = idx + 1
            break

    if status_col is None:
        wb.close()
        raise IssueListError("SS-IL-E-003", "问题清单无'状态'列")

    actual_row = row_index + 1  # +1 for header
    if actual_row > ws.max_row:
        wb.close()
        raise IssueListError("SS-IL-E-003", f"行号越界: {row_index}")

    ws.cell(row=actual_row, column=status_col, value=new_status.value)
    wb.save(issue_list_path)
    wb.close()


def get_open_issues_count(issue_list_path: str) -> int:
    """获取状态为 open 的问题数量。"""
    if not os.path.exists(issue_list_path):
        return 0

    wb = openpyxl.load_workbook(issue_list_path, read_only=True, data_only=True)
    ws = wb.active

    header_row = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    status_col = None
    for idx, h in enumerate(header_row):
        if h == "状态":
            status_col = idx
            break

    if status_col is None:
        wb.close()
        return 0

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        status = str(row[status_col] or "").strip()
        if status == IssueStatus.OPEN.value:
            count += 1

    wb.close()
    return count


# ──────────────────────────────────────────────
# 内部工具
# ──────────────────────────────────────────────

def _generate_issue(item: dict, llm_call: LLMCallable) -> IssueGenResult:
    """调用 LLM 生成单条问题描述。"""
    user_prompt = ISSUE_USER_TEMPLATE.format(
        check_content=item["check_content"],
        latest_result=item["latest_result"],
        assessment_conclusion=item["assessment"],
    )

    response = llm_call(ISSUE_SYSTEM_PROMPT, user_prompt)
    text = response.strip()

    if text.startswith("```"):
        lines = text.split("\n")
        text = "\n".join(lines[1:-1] if lines[-1].strip() == "```" else lines[1:])

    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        raise IssueListError("SS-IL-E-002", f"LLM 返回无法解析: {text[:200]}")

    sev = str(data.get("severity", "mid")).strip().lower()
    if sev not in _VALID_SEVERITY:
        sev = "mid"
    return IssueGenResult(
        problem_description=data.get("problem_description", ""),
        remediation_suggestion=data.get("remediation_suggestion", ""),
        severity=sev,
    )
