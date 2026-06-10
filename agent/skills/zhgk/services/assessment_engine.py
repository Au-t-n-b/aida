"""
智慧工勘 v4 — LLM 五值评估引擎

职责: 基于 LLM 对勘测结果做五值评估
错误码前缀: SS-AE
"""
from __future__ import annotations

import json
import os
from typing import Optional

import openpyxl

from .logger import log_error, log_info, log_warn
from .types import AssessmentResult, AssessmentValue, LLMCallable, SurveyResultRow


# ──────────────────────────────────────────────
# 异常定义
# ──────────────────────────────────────────────

class AssessmentError(Exception):
    def __init__(self, code: str, message: str):
        self.code = code
        self.message = message
        super().__init__(f"[{code}] {message}")


# ──────────────────────────────────────────────
# Prompt 模板
# ──────────────────────────────────────────────

ASSESSMENT_SYSTEM_PROMPT = """你是一名数据中心工勘质量评估专家。你的任务是根据「检查内容」（标准要求）和「检查结果」（实际勘测数据），
判断该勘测项是否满足要求。

你必须输出以下 5 种结论之一：
- 满足：检查结果完全符合检查内容的要求
- 不满足：检查结果明确不符合检查内容的要求
- 不涉及：该勘测项在当前场景中不适用（如液冷项目中的风冷专属检查）
- 无法识别：检查结果信息不足或模糊，无法做出明确判断

输出格式（严格 JSON）：
{
  "conclusion": "满足|不满足|不涉及|无法识别",
  "defect": "缺陷描述（仅在不满足或无法识别时填写，否则为空字符串）"
}

注意：
1. 只输出 JSON，不要加任何解释
2. "未勘测"由系统自动判断（检查结果为空时），你不需要输出此值
3. 判断应严格基于事实，不做主观推测"""

ASSESSMENT_USER_TEMPLATE = "检查内容: {check_content}\n检查结果: {latest_result}"


# ──────────────────────────────────────────────
# 有效评估值集合
# ──────────────────────────────────────────────

VALID_CONCLUSIONS = {"满足", "不满足", "不涉及", "无法识别"}


# ──────────────────────────────────────────────
# 公开接口
# ──────────────────────────────────────────────

def evaluate_single(
    check_content: str,
    latest_result: str,
    llm_call: LLMCallable,
) -> AssessmentResult:
    """
    对单条检查项做 LLM 评估。

    特殊逻辑:
        - latest_result 为空/None → 直接返回"未勘测"，不调用 LLM
    """
    if not latest_result or not latest_result.strip():
        return AssessmentResult(
            conclusion=AssessmentValue.NOT_SURVEYED,
            defect_description="",
            confidence=1.0,
        )

    user_prompt = ASSESSMENT_USER_TEMPLATE.format(
        check_content=check_content,
        latest_result=latest_result,
    )

    try:
        response = llm_call(ASSESSMENT_SYSTEM_PROMPT, user_prompt)
    except Exception as e:
        raise AssessmentError("SS-AE-E-001", f"LLM 调用失败: {e}")

    return _parse_response(response)


def evaluate_all(
    survey_table_path: str,
    llm_call: LLMCallable,
) -> list[AssessmentResult]:
    """
    批量评估全量勘测结果表中所有条目。

    副作用: 将评估结果写回 Excel 的"AI评估结果"列。
    错误处理: 单条失败记 WARNING 并标"无法识别"，继续下一行。
    """
    if not os.path.exists(survey_table_path):
        raise AssessmentError("SS-AE-E-003", f"文件不存在: {survey_table_path}")

    wb = openpyxl.load_workbook(survey_table_path)
    ws = wb.active

    # 定位列索引
    header_row = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    check_col = _find_col(header_row, "检查内容")
    result_col = _find_col(header_row, "最新检查结果")
    ai_col = _find_col(header_row, "AI评估结果")

    if check_col is None or result_col is None or ai_col is None:
        wb.close()
        raise AssessmentError(
            "SS-AE-E-003", "表格缺少必要列: 检查内容/最新检查结果/AI评估结果"
        )

    results: list[AssessmentResult] = []
    success_count = 0
    fail_count = 0

    for row_idx in range(2, ws.max_row + 1):
        seq_val = ws.cell(row_idx, 1).value
        if seq_val is None:
            break

        check_content = str(ws.cell(row_idx, check_col + 1).value or "").strip()
        latest_result = str(ws.cell(row_idx, result_col + 1).value or "").strip()

        try:
            ar = evaluate_single(check_content, latest_result, llm_call)
            ws.cell(row_idx, ai_col + 1, value=ar.conclusion.value)
            results.append(ar)
            success_count += 1
        except AssessmentError as e:
            log_warn("assessment_engine", "evaluate_all", f"行 {row_idx} 评估失败: {e.message}")
            ws.cell(row_idx, ai_col + 1, value="无法识别")
            results.append(AssessmentResult(
                conclusion=AssessmentValue.UNRECOGNIZABLE,
                defect_description=f"评估失败: {e.message}",
                confidence=0.0,
            ))
            fail_count += 1

    if success_count == 0 and fail_count > 0:
        wb.close()
        raise AssessmentError("SS-AE-E-001", "批量评估全部失败")

    wb.save(survey_table_path)
    wb.close()

    log_info("assessment_engine", "evaluate_all",
             f"评估完成: 成功 {success_count}, 失败 {fail_count}")
    return results


def get_assessment_statistics(survey_table_path: str) -> dict[str, int]:
    """
    统计各评估结果的数量。

    返回:
        {"满足": N, "不满足": N, "不涉及": N, "未勘测": N, "无法识别": N}
    """
    stats = {v.value: 0 for v in AssessmentValue}

    if not os.path.exists(survey_table_path):
        return stats

    wb = openpyxl.load_workbook(survey_table_path, read_only=True, data_only=True)
    ws = wb.active

    header_row = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    ai_col = _find_col(header_row, "AI评估结果")

    if ai_col is None:
        wb.close()
        return stats

    for row in ws.iter_rows(min_row=2, min_col=ai_col + 1, max_col=ai_col + 1, values_only=True):
        val = str(row[0] or "").strip()
        if val in stats:
            stats[val] += 1
        elif val == "":
            stats["未勘测"] += 1

    wb.close()
    return stats


# ──────────────────────────────────────────────
# 内部工具
# ──────────────────────────────────────────────

def _find_col(headers: list[str], name: str) -> Optional[int]:
    """在表头列表中找到指定列名的索引（0-based）。"""
    for idx, h in enumerate(headers):
        if h == name:
            return idx
    return None


def _parse_response(response: str) -> AssessmentResult:
    """解析 LLM 返回的 JSON 为 AssessmentResult。"""
    text = response.strip()
    # 去除可能的 markdown 代码块
    if text.startswith("```"):
        lines = text.split("\n")
        text = "\n".join(lines[1:-1] if lines[-1].strip() == "```" else lines[1:])

    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        raise AssessmentError("SS-AE-E-002", f"LLM 返回无法解析为 JSON: {text[:200]}")

    conclusion_str = data.get("conclusion", "")
    if conclusion_str not in VALID_CONCLUSIONS:
        raise AssessmentError(
            "SS-AE-E-002",
            f"LLM 返回无效结论值 '{conclusion_str}'，期望: {VALID_CONCLUSIONS}",
        )

    defect = data.get("defect", "")
    return AssessmentResult(
        conclusion=AssessmentValue.from_str(conclusion_str),
        defect_description=defect,
        confidence=0.9,
    )
