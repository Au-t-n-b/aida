"""
gkclaw ingest · 入站包处理（契约 §14/§17-§20 + 项目边界 A2/B5/B6/双源拍板）

路由：task.import_ack → accepted（一致性校验）；task.result → staged 只记录 /
final 转写已填写表走 wait_survey 原有合并通道；task.error → 留痕。
处置不抛栈（单包失败隔离，不阻断批次）；所有判定走 registry.decide_inbound 纯函数。

final 合并三道闸（任一不过 → 暂存 pending_results/ + 告警，merged=False）：
  1) submitted_by ∈ assignees（§22 安全基线，不过=整包隔离）
  2) 表指纹一致（下发后表未被 redo/追加）
  3) Input/ 无待合并的人工上传表（先到先得，用户拍板）
"""
from __future__ import annotations

import shutil
from pathlib import Path
from typing import Any

from . import package, schema
from .registry import TaskRegistry, decide_inbound

FILLED_TABLE_NAME = "已填写_全量勘测结果表.xlsx"


class MailboxExhausted(RuntimeError):
    """附件序号越界（探测附件数量用）。"""


def write_filled_table(survey_table_path: str, results: dict[int, str], dest_path: Path) -> int:
    """复制全量表并按序号写入「最新检查结果」列，存为已填写表。返回写入条数。"""
    import openpyxl

    wb = openpyxl.load_workbook(survey_table_path)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]
    seq_col = headers.index("序号") + 1
    res_col = headers.index("最新检查结果") + 1
    written = 0
    for row in ws.iter_rows(min_row=2):
        raw = row[seq_col - 1].value
        if raw is None:
            continue
        try:
            seq = int(raw)
        except (TypeError, ValueError):
            continue
        if seq in results:
            ws.cell(row[0].row, res_col, results[seq])
            written += 1
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest_path)
    wb.close()
    return written


def _quarantine(reg: TaskRegistry, zip_path: Path | str, note: str) -> dict[str, Any]:
    dest = reg.quarantine_dir() / Path(zip_path).name
    try:
        shutil.copy(zip_path, dest)
    except Exception:  # noqa: BLE001
        pass
    return {"disposition": "quarantine", "note": note, "merged": False}


def _apply_ack(reg: TaskRegistry, task: dict, payload: dict) -> tuple[str, str]:
    """ACK 一致性校验（§14）：项目/任务名/人员与原任务不一致 → 隔离。"""
    orig = reg.task_payload(task["task_id"]) or {}
    mismatches = []
    if payload.get("task_name") and payload["task_name"] != orig.get("task_name"):
        mismatches.append("task_name")
    if (payload.get("project") or {}).get("project_code") != (orig.get("project") or {}).get("project_code"):
        mismatches.append("project_code")
    orig_codes = {a.get("surveyor_code") for a in orig.get("assignees", [])}
    ack_codes = {a.get("surveyor_code") for a in payload.get("assignees", [])}
    if ack_codes and ack_codes != orig_codes:
        mismatches.append("assignees")
    if mismatches:
        return "quarantine", f"ACK 与原任务不一致（{'/'.join(mismatches)}），隔离待人工核查"
    reg.set_state(task["task_id"], "accepted",
                  web_access_url=str(payload.get("web_access_url", "")),
                  accepted_at=str(payload.get("accepted_at", "")))
    return "processed", "frontagent 已导入任务"


def _apply_result(
    reg: TaskRegistry, task: dict, payload: dict, zip_path: Path,
    *, input_dir: Path, survey_table_path: str | None,
) -> dict[str, Any]:
    tid = task["task_id"]
    # 安全闸 1：submitted_by ∈ assignees（隔离，不入库）
    code = str((payload.get("submitted_by") or {}).get("surveyor_code", ""))
    allowed = {a.get("surveyor_code") for a in task.get("assignees", [])}
    if code not in allowed:
        return _quarantine(reg, zip_path, f"submitted_by({code}) 不在任务 assignees 中，整包隔离")

    is_final = (payload.get("session") or {}).get("status") == "completed"
    reg.store_result_version(tid, payload, final=is_final)

    # evidence 解出 + to_back_备注 留档
    parsed_names = [n for n in package.parse_package(zip_path)["files"]
                    if n.startswith("evidence/") and not n.endswith("/")]
    if parsed_names:
        package.extract_files(zip_path, parsed_names, reg.root / tid)
    notes = [{"问题序号": str(it.get("问题序号", "")), "to_back_备注": it["to_back_备注"]}
             for it in payload.get("items", []) if str(it.get("to_back_备注", "")).strip()]
    if notes:
        reg.append_result_notes(tid, notes)

    if not is_final:
        reg.set_state(tid, "staged_returned")
        return {"disposition": "processed", "merged": False,
                "note": "阶段性回传已记录（不推进流程，final 才合并）"}

    # final：任务完成；合并需过闸 2/3
    reg.set_state(tid, "completed")
    known_keys = {it["问题序号"] for it in (reg.task_payload(tid) or {}).get("items", [])}
    results: dict[int, str] = {}
    skipped: list[str] = []
    for it in payload.get("items", []):
        key = str(it.get("问题序号", ""))
        val = str(it.get("勘测结果", "")).strip()
        if not val:
            continue
        if key not in known_keys or not key.isdigit():
            skipped.append(key)
            continue
        results[int(key)] = val

    def _stash(reason: str) -> dict[str, Any]:
        pend = reg.root / tid / "pending_results"
        pend.mkdir(parents=True, exist_ok=True)
        if survey_table_path:
            write_filled_table(survey_table_path, results, pend / FILLED_TABLE_NAME)
        else:
            shutil.copy(zip_path, pend / Path(zip_path).name)
        reg.set_state(tid, "completed", merge_blocked=True, merge_blocked_reason=reason)
        return {"disposition": "processed", "merged": False, "note": reason}

    if not survey_table_path:
        return _stash("找不到当前全量勘测结果表，结果已暂存 pending_results/ 待人工合并")
    fingerprint = package.sha256_file(survey_table_path)
    if fingerprint != task.get("table_fingerprint"):
        return _stash("表指纹不一致（下发后勘测表已变更），结果暂存 pending_results/ 待人工裁决")
    dest = Path(input_dir) / FILLED_TABLE_NAME
    if dest.exists():
        return _stash("先到先得：Input/ 已有待合并的人工上传表，邮件结果暂存 pending_results/")
    written = write_filled_table(survey_table_path, results, dest)
    note = f"最终回传已转写 {written} 条到 Input/{FILLED_TABLE_NAME}（走 wait_survey 合并通道）"
    if skipped:
        note += f"；跳过未下发条目 {skipped}"
    return {"disposition": "processed", "merged": True, "note": note}


def ingest_zip(
    zip_path: Path | str,
    *,
    runtime_dir: Path | str,
    input_dir: Path | str,
    survey_table_path: str | None,
    mail_id: int | None = None,
) -> dict[str, Any]:
    """处理单个入站 ZIP。永不抛栈；返回 {disposition, note, merged, task_id?}。"""
    reg = TaskRegistry(runtime_dir)
    try:
        parsed = package.parse_package(zip_path)
        if not parsed["ok"]:
            return _quarantine(reg, zip_path, "包校验失败: " + "；".join(parsed["errors"]))
        manifest = parsed["manifest"]
        payload = parsed["payload"]
        ptype = manifest["package_type"]
        validator = {"task.import_ack": schema.validate_ack,
                     "task.result": schema.validate_result,
                     "task.error": schema.validate_error}.get(ptype)
        if validator is None:
            return _quarantine(reg, zip_path, f"backagent 不接收 {ptype} 包")
        errors = validator(payload)
        if errors:
            return _quarantine(reg, zip_path, "payload 校验失败: " + "；".join(errors))

        task_id = manifest["task_id"]
        task = reg.get(task_id)
        checksum = package.sha256_file(zip_path)
        session_status = str((payload.get("session") or {}).get("status", "")) \
            if ptype == "task.result" else ""
        disposition, note = decide_inbound(
            task, reg.packages(task_id) if task else [],
            package_type=ptype, package_id=manifest["package_id"],
            checksum=checksum, session_status=session_status,
        )
        out: dict[str, Any] = {"disposition": disposition, "note": note,
                               "merged": False, "task_id": task_id}
        if disposition == "processed":
            if ptype == "task.import_ack":
                disposition, note = _apply_ack(reg, task, payload)
                out.update({"disposition": disposition, "note": note})
                if disposition == "quarantine":
                    _quarantine(reg, zip_path, note)
            elif ptype == "task.result":
                out.update(_apply_result(reg, task, payload, Path(zip_path),
                                         input_dir=Path(input_dir),
                                         survey_table_path=survey_table_path))
            elif ptype == "task.error":
                recoverable = bool(payload.get("recoverable", False))
                err = f"[{payload.get('code')}] {payload.get('message')}"
                if recoverable:
                    reg.set_state(task_id, "dispatched", last_error=err)
                    out["note"] = f"frontagent 报可恢复错误：{err}（修复后用新 task_id 重发）"
                else:
                    reg.set_state(task_id, "failed", last_error=err)
                    out["note"] = f"frontagent 报不可恢复错误：{err}"
        elif disposition in ("conflict", "quarantine"):
            _quarantine(reg, zip_path, note)
        if task is not None:
            reg.record_package(task_id, {
                "package_id": manifest["package_id"], "checksum": checksum,
                "package_type": ptype, "direction": "in",
                "disposition": out["disposition"], "mail_id": mail_id,
            })
        return out
    except Exception as e:  # noqa: BLE001 — 单包异常隔离，不阻断批次
        return _quarantine(reg, zip_path, f"处理异常: {type(e).__name__}: {e}")


def poll_and_ingest(
    *,
    runtime_dir: Path | str,
    input_dir: Path | str,
    survey_table_path: str | None,
    mailbox_mod: Any = None,
    limit: int = 50,
    max_attachments: int = 10,
) -> dict[str, Any]:
    """拉取 mailgw 收件箱并处理 GKCLAW 包。

    边界 B5：只列表+另存附件（不读正文/不动已读）；mail_id 账本去重；
    非 ZIP/非 GKCLAW 邮件记 ignored 不再重扫。返回 {checked, processed, alerts[]}。
    """
    if mailbox_mod is None:
        import agent.mailbox as mailbox_mod  # type: ignore[no-redef]
    reg = TaskRegistry(runtime_dir)
    summary: dict[str, Any] = {"checked": 0, "processed": 0, "alerts": []}
    if not mailbox_mod.is_configured():
        summary["alerts"].append("mailgw 未配置（MAILGW_TOKEN），跳过邮件拉取")
        return summary

    try:
        mails = mailbox_mod.list_inbox(refresh=True, limit=limit, unread_only=False)["mails"]
    except Exception as e:  # noqa: BLE001
        summary["alerts"].append(f"收件箱拉取失败: {e}")
        return summary

    scanned = reg.scanned_mail_ids()
    for mail in mails:
        mail_id = int(mail.get("mail_id", -1))
        if mail_id in scanned:
            continue
        summary["checked"] += 1
        if not mail.get("has_attachments"):
            reg.mark_mail_scanned(mail_id, "ignored")
            continue
        tmp = reg.root / "_inbox_tmp" / str(mail_id)
        tmp.mkdir(parents=True, exist_ok=True)
        verdict = "ignored"
        for idx in range(max_attachments):
            try:
                saved = mailbox_mod.save_attachment(mail_id, idx, str(tmp))
            except Exception:  # noqa: BLE001 — 序号越界/网关错误 → 该邮件附件取尽
                break
            if not saved.lower().endswith(".zip"):
                continue
            out = ingest_zip(saved, runtime_dir=runtime_dir, input_dir=input_dir,
                             survey_table_path=survey_table_path, mail_id=mail_id)
            verdict = out["disposition"]
            if out["disposition"] == "processed":
                summary["processed"] += 1
            if out.get("note"):
                summary["alerts"].append(f"mail#{mail_id}: {out['note']}")
        reg.mark_mail_scanned(mail_id, verdict)
    return summary
