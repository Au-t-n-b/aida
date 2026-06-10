"""
智慧工勘 v4 — 全量勘测结果表构建

职责: 创建全量勘测结果表、追加数据条目、追加自定义条目
错误码前缀: SS-ST
"""
from __future__ import annotations

import os
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter

from .types import SurveyItem, SurveyResultRow


# ──────────────────────────────────────────────
# 异常定义
# ──────────────────────────────────────────────

class SurveyTableError(Exception):
    def __init__(self, code: str, message: str):
        self.code = code
        self.message = message
        super().__init__(f"[{code}] {message}")


# ──────────────────────────────────────────────
# 常量
# ──────────────────────────────────────────────

RESULT_TABLE_HEADERS = [
    "序号", "细分场景", "勘测要素", "项目", "检查内容",
    "勘测方法", "最新检查结果", "AI评估结果",
    "图片1", "图片2", "第一轮勘测结果", "备注",
]

# 从 SurveyItem 到 SurveyResultRow 的列映射
COLUMN_MAPPING = {
    "细分场景": "细分场景",
    "勘测要素": "勘测要素",
    "项目": "项目",
    "检查内容": "检查内容",
    "勘测方法": "勘测方法",
    "备注": "备注",
}


# ──────────────────────────────────────────────
# 公开接口
# ──────────────────────────────────────────────

def build_survey_table(
    filtered_items: list[SurveyItem],
    output_dir: str,
    activity_id: str,
    project_name: str,
    room_name: str,
) -> str:
    """
    从过滤后的底表条目构建全量勘测结果表。

    移除元数据列（代际-制冷、分类、是否支持视频勘测、背景知识），
    保留细分场景，新增功能列（AI评估结果、图片、勘测结果列）。

    返回:
        生成的文件完整路径
    """
    parts = [p for p in [activity_id, project_name, room_name] if p]
    prefix = "_".join(parts)
    filename = f"{prefix}_全量勘测结果表.xlsx" if prefix else "全量勘测结果表.xlsx"
    output_path = os.path.join(output_dir, filename)

    os.makedirs(output_dir, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "入场评估标准"

    # 写入表头
    for col_idx, header in enumerate(RESULT_TABLE_HEADERS, 1):
        ws.cell(row=1, column=col_idx, value=header)

    # 写入数据行
    for row_idx, item in enumerate(filtered_items, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)  # 序号
        ws.cell(row=row_idx, column=2, value=item.get("细分场景", ""))
        ws.cell(row=row_idx, column=3, value=item.get("勘测要素", ""))
        ws.cell(row=row_idx, column=4, value=item.get("项目", ""))
        ws.cell(row=row_idx, column=5, value=item.get("检查内容", ""))
        ws.cell(row=row_idx, column=6, value=item.get("勘测方法", ""))
        ws.cell(row=row_idx, column=7, value="")  # 最新检查结果
        ws.cell(row=row_idx, column=8, value="")  # AI评估结果
        ws.cell(row=row_idx, column=9, value=None)  # 图片1
        ws.cell(row=row_idx, column=10, value=None)  # 图片2
        ws.cell(row=row_idx, column=11, value="")  # 第一轮勘测结果
        ws.cell(row=row_idx, column=12, value=item.get("备注", ""))

    # 添加"要求" Sheet
    ws_req = wb.create_sheet("要求")
    ws_req.cell(row=1, column=1, value="业务需求备注")
    ws_req.cell(row=2, column=1, value="需求1：辅料采购逻辑（接地线/定位螺栓/排水管/电源线）")
    ws_req.cell(row=3, column=1, value="需求2：支持项目组自建标准 + AI辅助判断")

    wb.save(output_path)
    wb.close()
    return output_path


def append_data_items(
    survey_table_path: str,
    data_items: list[SurveyItem],
) -> int:
    """
    追加"分类=数据"的条目到已有全量勘测结果表。
    序号自动接续当前最大值。

    返回:
        追加后的总行数
    """
    if not os.path.exists(survey_table_path):
        raise SurveyTableError("SS-ST-E-001", f"文件不存在: {survey_table_path}")

    wb = openpyxl.load_workbook(survey_table_path)
    ws = wb.active

    max_seq = get_max_sequence(survey_table_path=None, ws=ws)

    for i, item in enumerate(data_items, 1):
        new_seq = max_seq + i
        row_idx = ws.max_row + 1
        ws.cell(row=row_idx, column=1, value=new_seq)
        ws.cell(row=row_idx, column=2, value=item.get("细分场景", ""))
        ws.cell(row=row_idx, column=3, value=item.get("勘测要素", ""))
        ws.cell(row=row_idx, column=4, value=item.get("项目", ""))
        ws.cell(row=row_idx, column=5, value=item.get("检查内容", ""))
        ws.cell(row=row_idx, column=6, value=item.get("勘测方法", ""))
        ws.cell(row=row_idx, column=7, value="")  # 最新检查结果
        ws.cell(row=row_idx, column=8, value="")  # AI评估结果
        ws.cell(row=row_idx, column=9, value=None)  # 图片1
        ws.cell(row=row_idx, column=10, value=None)  # 图片2
        ws.cell(row=row_idx, column=11, value="")  # 第一轮勘测结果
        ws.cell(row=row_idx, column=12, value=item.get("备注", ""))

    wb.save(survey_table_path)
    wb.close()
    return ws.max_row - 1  # 减去表头行


def append_custom_item(
    survey_table_path: str,
    item: dict,
) -> int:
    """
    追加用户自定义条目（底表中不存在的全新勘测项）。

    参数:
        item: 必须包含 {"细分场景", "勘测要素", "项目", "检查内容", "勘测方法"}

    返回:
        新条目的序号
    """
    if not os.path.exists(survey_table_path):
        raise SurveyTableError("SS-ST-E-001", f"文件不存在: {survey_table_path}")

    required_keys = {"细分场景", "勘测要素", "项目", "检查内容", "勘测方法"}
    missing = required_keys - set(item.keys())
    if missing:
        raise SurveyTableError("SS-ST-E-002", f"自定义条目缺少必填字段: {missing}")

    wb = openpyxl.load_workbook(survey_table_path)
    ws = wb.active

    max_seq = get_max_sequence(survey_table_path=None, ws=ws)
    new_seq = max_seq + 1
    row_idx = ws.max_row + 1

    ws.cell(row=row_idx, column=1, value=new_seq)
    ws.cell(row=row_idx, column=2, value=item.get("细分场景", ""))
    ws.cell(row=row_idx, column=3, value=item.get("勘测要素", ""))
    ws.cell(row=row_idx, column=4, value=item.get("项目", ""))
    ws.cell(row=row_idx, column=5, value=item.get("检查内容", ""))
    ws.cell(row=row_idx, column=6, value=item.get("勘测方法", ""))
    ws.cell(row=row_idx, column=7, value="")
    ws.cell(row=row_idx, column=8, value="")
    ws.cell(row=row_idx, column=9, value=None)
    ws.cell(row=row_idx, column=10, value=None)
    ws.cell(row=row_idx, column=11, value="")
    ws.cell(row=row_idx, column=12, value=item.get("备注", ""))

    wb.save(survey_table_path)
    wb.close()
    return new_seq


def read_survey_table(survey_table_path: str) -> list[SurveyResultRow]:
    """
    读取全量勘测结果表的全部数据行。
    按列名取值，动态识别轮次列。
    """
    if not os.path.exists(survey_table_path):
        raise SurveyTableError("SS-ST-E-001", f"文件不存在: {survey_table_path}")

    wb = openpyxl.load_workbook(survey_table_path, read_only=True, data_only=True)
    ws = wb.active

    # 读取表头
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h).strip() if h else "" for h in header_row]
    col_map = {name: idx for idx, name in enumerate(headers) if name}

    rows: list[SurveyResultRow] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        result_row: SurveyResultRow = {
            "序号": _safe_int(row, col_map.get("序号")),
            "细分场景": _safe_str(row, col_map.get("细分场景")),
            "勘测要素": _safe_str(row, col_map.get("勘测要素")),
            "项目": _safe_str(row, col_map.get("项目")),
            "检查内容": _safe_str(row, col_map.get("检查内容")),
            "勘测方法": _safe_str(row, col_map.get("勘测方法")),
            "最新检查结果": _safe_str(row, col_map.get("最新检查结果")),
            "AI评估结果": _safe_str(row, col_map.get("AI评估结果")),
            "图片1": None,
            "图片2": None,
            "备注": _safe_str(row, col_map.get("备注")),
        }
        rows.append(result_row)

    wb.close()
    return rows


def get_max_sequence(
    survey_table_path: Optional[str] = None,
    ws=None,
) -> int:
    """获取当前表中最大序号值。"""
    if ws is None:
        if survey_table_path is None:
            return 0
        wb = openpyxl.load_workbook(survey_table_path, read_only=True, data_only=True)
        ws = wb.active
        close_after = True
    else:
        close_after = False

    max_seq = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        val = row[0]
        if val is not None:
            try:
                seq = int(val)
                if seq > max_seq:
                    max_seq = seq
            except (ValueError, TypeError):
                pass

    if close_after:
        wb.close()
    return max_seq


# ──────────────────────────────────────────────
# 内部工具
# ──────────────────────────────────────────────

def _safe_str(row: tuple, idx: Optional[int]) -> str:
    if idx is None or idx >= len(row):
        return ""
    val = row[idx]
    return str(val).strip() if val is not None else ""


def _safe_int(row: tuple, idx: Optional[int]) -> int:
    if idx is None or idx >= len(row):
        return 0
    val = row[idx]
    if val is None:
        return 0
    try:
        return int(val)
    except (ValueError, TypeError):
        return 0
