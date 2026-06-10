"""
智慧工勘 v4 — 风险识别引擎

职责: 基于高风险库和勘测结果做 LLM 风险判断
错误码前缀: SS-RE
"""
from __future__ import annotations

import json
import os
from typing import Optional

import openpyxl

from .logger import log_info, log_warn
from .types import LLMCallable, RiskItem, RiskJudgment, SurveyResultRow


# ──────────────────────────────────────────────
# 异常定义
# ──────────────────────────────────────────────

class RiskEngineError(Exception):
    def __init__(self, code: str, message: str):
        self.code = code
        self.message = message
        super().__init__(f"[{code}] {message}")


# ──────────────────────────────────────────────
# Prompt 模板
# ──────────────────────────────────────────────

RISK_SYSTEM_PROMPT = """你是一名数据中心风险评估专家。根据「风险描述」和「相关勘测结果」，
判断该风险在当前项目中是否被触发。

输出格式（严格 JSON）：
{
  "triggered": true或false,
  "risk_impact": "如果触发，描述风险可能造成的影响（未触发则为空）",
  "suggestion": "如果触发，建议的应对措施（未触发则为空）"
}

注意：
1. 只输出 JSON
2. 判断应保守：有证据表明风险存在则判定为触发
3. 如果相关勘测结果不足以判断，也视为触发（预防性原则）"""

RISK_USER_TEMPLATE = (
    "风险描述: {risk_description}\n"
    "风险细项: {risk_detail}\n"
    "触发条件: {trigger_condition}\n"
    "相关勘测结果:\n{relevant_results}"
)


# ──────────────────────────────────────────────
# 高风险库列名
# ──────────────────────────────────────────────

COL_SEQ = "序号"
COL_RISK_TYPE = "风险类型描述"
COL_RISK_DETAIL = "风险细项描述"
COL_TRIGGER = "触发条件"
COL_LEVEL = "风险等级"
COL_RELATED_SURVEY = "关联勘测项"
COL_RELATED_ASSESS = "关联评估项"
COL_IMPACT = "风险影响"
COL_SUGGESTION = "风险应对建议"
COL_SCENE = "场景"
COL_KNOWLEDGE = "背景知识灌入"
COL_ENABLED = "是否启用"


# ──────────────────────────────────────────────
# 风险结果表头
# ──────────────────────────────────────────────

RISK_TABLE_HEADERS = [
    "序号", "风险类型描述", "风险细项描述", "风险影响", "建议措施", "备注",
]


# ──────────────────────────────────────────────
# 公开接口
# ──────────────────────────────────────────────

def load_risk_library(
    risk_library_path: str,
    generation_cooling: Optional[str] = None,
) -> list[dict]:
    """
    加载高风险库并按启用状态过滤。

    返回:
        启用的风险条目列表
    """
    if not os.path.exists(risk_library_path):
        raise RiskEngineError("SS-RE-E-001", f"高风险库文件不存在: {risk_library_path}")

    wb = openpyxl.load_workbook(risk_library_path, read_only=True, data_only=True)
    ws = wb.active

    # 读取表头
    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    col_map = {name: idx for idx, name in enumerate(headers) if name}

    # 找到"是否启用"列（可能带括号后缀）
    enabled_col = None
    for h, idx in col_map.items():
        if "是否启用" in h:
            enabled_col = idx
            break

    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue

        # 过滤启用状态
        if enabled_col is not None:
            enabled_val = row[enabled_col]
            if enabled_val not in (1, "1", True):
                continue

        item = {
            "序号": _safe_val(row, col_map.get(COL_SEQ)),
            "风险类型描述": _safe_str(row, col_map.get(COL_RISK_TYPE)),
            "风险细项描述": _safe_str(row, col_map.get(COL_RISK_DETAIL)),
            "触发条件": _safe_str(row, col_map.get(COL_TRIGGER)),
            "风险等级": _safe_str(row, col_map.get(COL_LEVEL)),
            "关联勘测项": _safe_str(row, col_map.get(COL_RELATED_SURVEY)),
            "关联评估项": _safe_str(row, col_map.get(COL_RELATED_ASSESS)),
            "风险影响": _safe_str(row, col_map.get(COL_IMPACT)),
            "风险应对建议": _safe_str(row, col_map.get(COL_SUGGESTION)),
            "场景": _safe_str(row, col_map.get(COL_SCENE)),
            "背景知识": _safe_str(row, col_map.get(COL_KNOWLEDGE)),
        }
        items.append(item)

    wb.close()
    return items


def identify_risks(
    risk_items: list[dict],
    survey_results: list[SurveyResultRow],
    llm_call: LLMCallable,
) -> list[RiskItem]:
    """
    逐条风险结合勘测结果做 LLM 判断。

    返回:
        已触发的 RiskItem 列表
    """
    triggered_risks: list[RiskItem] = []

    for i, risk in enumerate(risk_items):
        # 收集相关勘测结果
        relevant = _find_relevant_results(risk, survey_results)
        relevant_text = _format_relevant_results(relevant)

        user_prompt = RISK_USER_TEMPLATE.format(
            risk_description=risk.get("风险类型描述", ""),
            risk_detail=risk.get("风险细项描述", ""),
            trigger_condition=risk.get("触发条件", ""),
            relevant_results=relevant_text if relevant_text else "（无直接关联的勘测结果）",
        )

        try:
            response = llm_call(RISK_SYSTEM_PROMPT, user_prompt)
            judgment = _parse_risk_response(response)
        except Exception as e:
            log_warn("risk_engine", "identify_risks", f"风险 {i+1} LLM判断失败: {e}")
            # 预防性原则：判断失败视为触发
            judgment = RiskJudgment(
                triggered=True,
                risk_impact=risk.get("风险影响", ""),
                suggestion=risk.get("风险应对建议", ""),
            )

        if judgment.triggered:
            triggered_risks.append(RiskItem(
                序号=len(triggered_risks) + 1,
                风险类型描述=risk.get("风险类型描述", ""),
                风险细项描述=risk.get("风险细项描述", ""),
                风险影响=judgment.risk_impact or risk.get("风险影响", ""),
                建议措施=judgment.suggestion or risk.get("风险应对建议", ""),
                备注="",
            ))

    log_info("risk_engine", "identify_risks",
             f"风险判断完成: {len(risk_items)} 条检查, {len(triggered_risks)} 条触发")
    return triggered_risks


def write_risk_table(
    risks: list[RiskItem],
    output_dir: str,
    activity_id: str,
    project_name: str,
    room_name: str,
) -> str:
    """将风险列表写入风险识别结果表 Excel。"""
    os.makedirs(output_dir, exist_ok=True)
    filename = f"{activity_id}_{project_name}_{room_name}_风险识别结果表.xlsx"
    output_path = os.path.join(output_dir, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "风险识别结果"

    for col_idx, header in enumerate(RISK_TABLE_HEADERS, 1):
        ws.cell(row=1, column=col_idx, value=header)

    for i, risk in enumerate(risks, 1):
        ws.cell(i + 1, 1, value=i)
        ws.cell(i + 1, 2, value=risk["风险类型描述"])
        ws.cell(i + 1, 3, value=risk["风险细项描述"])
        ws.cell(i + 1, 4, value=risk["风险影响"])
        ws.cell(i + 1, 5, value=risk["建议措施"])
        ws.cell(i + 1, 6, value=risk.get("备注", ""))

    wb.save(output_path)
    wb.close()

    log_info("risk_engine", "write_risk_table", f"风险表已生成: {len(risks)} 条 → {output_path}")
    return output_path


# ──────────────────────────────────────────────
# 内部工具
# ──────────────────────────────────────────────

def _find_relevant_results(
    risk: dict,
    survey_results: list[SurveyResultRow],
) -> list[SurveyResultRow]:
    """根据"关联勘测项"或关键词匹配找到相关勘测结果。"""
    related_str = risk.get("关联勘测项", "")

    # 尝试解析关联行号（如 "48", "43,45"）
    if related_str:
        try:
            seq_nums = [int(s.strip()) for s in related_str.split(",") if s.strip()]
            matched = [r for r in survey_results if r.get("序号") in seq_nums]
            if matched:
                return matched
        except ValueError:
            pass

    # 回退：基于风险描述关键词匹配
    risk_type = risk.get("风险类型描述", "")
    risk_detail = risk.get("风险细项描述", "")
    keywords = _extract_keywords(risk_type + " " + risk_detail)

    if not keywords:
        return survey_results[:5]  # 无法匹配则取前5条作为上下文

    matched = []
    for r in survey_results:
        content = r.get("检查内容", "") + r.get("项目", "")
        if any(kw in content for kw in keywords):
            matched.append(r)

    return matched[:10] if matched else survey_results[:5]


def _extract_keywords(text: str) -> list[str]:
    """从风险描述中提取关键词（简单分词）。"""
    keywords = []
    for word in ["承重", "温度", "湿度", "配电", "UPS", "消防", "走线",
                 "制冷", "CDU", "管路", "漏水", "接地", "液冷", "风冷",
                 "排水", "加电", "通液"]:
        if word in text:
            keywords.append(word)
    return keywords


def _format_relevant_results(results: list[SurveyResultRow]) -> str:
    """格式化相关勘测结果为文本。"""
    if not results:
        return ""
    lines = []
    for r in results:
        check = r.get("检查内容", "")
        result = r.get("最新检查结果", "")
        ai = r.get("AI评估结果", "")
        if result:
            lines.append(f"- {check} → 结果: {result} (评估: {ai})")
        else:
            lines.append(f"- {check} → 未勘测")
    return "\n".join(lines)


def _parse_risk_response(response: str) -> RiskJudgment:
    """解析 LLM 风险判断响应。"""
    text = response.strip()
    if text.startswith("```"):
        lines = text.split("\n")
        text = "\n".join(lines[1:-1] if lines[-1].strip() == "```" else lines[1:])

    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        raise RiskEngineError("SS-RE-E-002", f"LLM 返回无法解析: {text[:200]}")

    triggered = data.get("triggered", False)
    if isinstance(triggered, str):
        triggered = triggered.lower() in ("true", "1", "yes")

    return RiskJudgment(
        triggered=triggered,
        risk_impact=data.get("risk_impact", ""),
        suggestion=data.get("suggestion", ""),
    )


def _safe_str(row: tuple, idx: Optional[int]) -> str:
    if idx is None or idx >= len(row):
        return ""
    val = row[idx]
    return str(val).strip() if val is not None else ""


def _safe_val(row: tuple, idx: Optional[int]):
    if idx is None or idx >= len(row):
        return None
    return row[idx]
