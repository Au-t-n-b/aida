"""
智慧工勘 v4 — 复勘管理

职责: 管理多轮勘测的结果写入和动态列扩展
错误码前缀: SS-RM
"""
from __future__ import annotations

import os
import re
from typing import Optional

import openpyxl

from .logger import log_info, log_warn
from .types import SurveyResultRow


# ──────────────────────────────────────────────
# 异常定义
# ──────────────────────────────────────────────

class ResurveyError(Exception):
    def __init__(self, code: str, message: str):
        self.code = code
        self.message = message
        super().__init__(f"[{code}] {message}")


# ──────────────────────────────────────────────
# 工具函数
# ──────────────────────────────────────────────

CHINESE_NUMBERS = ["一", "二", "三", "四", "五", "六", "七", "八", "九", "十"]

ROUND_PATTERN = re.compile(r"第(.+)轮勘测结果")


def to_chinese_number(n: int) -> str:
    """将阿拉伯数字转换为中文数字（1→一, 2→二, ...）"""
    if 1 <= n <= 10:
        return CHINESE_NUMBERS[n - 1]
    return str(n)


def from_chinese_number(s: str) -> int:
    """中文数字转阿拉伯数字。"""
    if s in CHINESE_NUMBERS:
        return CHINESE_NUMBERS.index(s) + 1
    try:
        return int(s)
    except (ValueError, TypeError):
        return 0


# ──────────────────────────────────────────────
# 公开接口
# ──────────────────────────────────────────────

def get_current_round(survey_table_path: str) -> int:
    """
    检测当前全量勘测结果表已有多少轮勘测结果列。

    返回:
        下一轮的轮次号（如已有"第一轮"和"第二轮"则返回 3）。
        无轮次列时返回 1。
    """
    if not os.path.exists(survey_table_path):
        raise ResurveyError("SS-RM-E-002", f"文件不存在: {survey_table_path}")

    wb = openpyxl.load_workbook(survey_table_path, read_only=True, data_only=True)
    ws = wb.active

    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    wb.close()

    max_round = 0
    for h in headers:
        m = ROUND_PATTERN.match(h)
        if m:
            n = from_chinese_number(m.group(1))
            if n > max_round:
                max_round = n

    return max_round + 1


def write_survey_results(
    survey_table_path: str,
    results: dict[int, str],
    round_number: int,
) -> None:
    """
    写入本轮勘测结果。

    参数:
        survey_table_path: 全量勘测结果表路径
        results: {序号(1-based): 勘测结果文本} 字典
        round_number: 当前轮次号

    副作用:
        1. 覆盖 "最新检查结果" 列
        2. 存档到 "第{round_number}轮勘测结果" 列
        3. 如果该轮次列不存在则动态新增
    """
    if not os.path.exists(survey_table_path):
        raise ResurveyError("SS-RM-E-002", f"文件不存在: {survey_table_path}")

    wb = openpyxl.load_workbook(survey_table_path)
    ws = wb.active

    # 获取表头
    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]

    # 定位"最新检查结果"列
    latest_col = None
    for idx, h in enumerate(headers):
        if h == "最新检查结果":
            latest_col = idx + 1  # openpyxl 1-based
            break

    if latest_col is None:
        wb.close()
        raise ResurveyError("SS-RM-E-001", "表格缺少'最新检查结果'列")

    # 确定轮次列名
    round_col_name = f"第{to_chinese_number(round_number)}轮勘测结果"

    # 检查轮次列是否存在，不存在则新增
    round_col = None
    for idx, h in enumerate(headers):
        if h == round_col_name:
            round_col = idx + 1
            break

    if round_col is None:
        # 找到最后一个轮次列的位置，在其后插入
        last_round_col = _find_last_round_col(headers)
        if last_round_col is not None:
            insert_pos = last_round_col + 1
        else:
            # 没有任何轮次列，在"AI评估结果"之后或表末插入
            ai_col = None
            for idx, h in enumerate(headers):
                if h == "AI评估结果":
                    ai_col = idx + 1
                    break
            insert_pos = (ai_col + 1) if ai_col else ws.max_column + 1

        # 在 insert_pos 位置插入新列
        ws.insert_cols(insert_pos)
        ws.cell(1, insert_pos, value=round_col_name)
        round_col = insert_pos

        # 重新读取 headers（插入后索引变化）
        headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
        for idx, h in enumerate(headers):
            if h == "最新检查结果":
                latest_col = idx + 1
                break

    # 建立序号→行号映射
    seq_to_row = {}
    for row_idx in range(2, ws.max_row + 1):
        seq_val = ws.cell(row_idx, 1).value
        if seq_val is not None:
            try:
                seq_to_row[int(seq_val)] = row_idx
            except (ValueError, TypeError):
                pass

    # 写入结果
    for seq, value in results.items():
        row_idx = seq_to_row.get(seq)
        if row_idx is None:
            log_warn("resurvey_manager", "write_survey_results",
                     f"序号 {seq} 未找到对应行，已跳过")
            continue
        ws.cell(row_idx, latest_col, value=value)
        ws.cell(row_idx, round_col, value=value)

    wb.save(survey_table_path)
    wb.close()

    log_info("resurvey_manager", "write_survey_results",
             f"第{to_chinese_number(round_number)}轮结果已写入 {len(results)} 条")


def get_latest_result(row: SurveyResultRow) -> str:
    """直接读取"最新检查结果"的值。"""
    return row.get("最新检查结果", "")


def get_items_needing_resurvey(survey_table_path: str) -> list[dict]:
    """
    获取需要复勘的条目列表。

    筛选条件: AI评估结果 in ["不满足", "无法识别", "未勘测"]
    """
    RESURVEY_TRIGGERS = {"不满足", "无法识别", "未勘测"}

    if not os.path.exists(survey_table_path):
        raise ResurveyError("SS-RM-E-002", f"文件不存在: {survey_table_path}")

    wb = openpyxl.load_workbook(survey_table_path, read_only=True, data_only=True)
    ws = wb.active

    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    col_map = {name: idx for idx, name in enumerate(headers) if name}

    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue

        ai_val = str(row[col_map.get("AI评估结果", -1)] or "").strip() if "AI评估结果" in col_map else ""

        if ai_val in RESURVEY_TRIGGERS:
            items.append({
                "序号": int(row[0]) if row[0] else 0,
                "项目": str(row[col_map.get("项目", -1)] or "").strip() if "项目" in col_map else "",
                "检查内容": str(row[col_map.get("检查内容", -1)] or "").strip() if "检查内容" in col_map else "",
                "AI评估结果": ai_val,
                "最新检查结果": str(row[col_map.get("最新检查结果", -1)] or "").strip() if "最新检查结果" in col_map else "",
            })

    wb.close()
    return items


def get_round_history(survey_table_path: str, row_index: int) -> dict[str, str]:
    """
    获取某行的所有轮次历史。

    参数:
        row_index: 序号（1-based 数据行序号）

    返回:
        {"第一轮勘测结果": "...", "第二轮勘测结果": "...", ...}
    """
    if not os.path.exists(survey_table_path):
        raise ResurveyError("SS-RM-E-002", f"文件不存在: {survey_table_path}")

    wb = openpyxl.load_workbook(survey_table_path, read_only=True, data_only=True)
    ws = wb.active

    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]

    # 找到目标行
    target_row = None
    for r in range(2, ws.max_row + 1):
        seq_val = ws.cell(r, 1).value
        if seq_val is not None and int(seq_val) == row_index:
            target_row = r
            break

    if target_row is None:
        wb.close()
        raise ResurveyError("SS-RM-E-003", f"序号 {row_index} 未找到")

    history = {}
    for col_idx, h in enumerate(headers):
        if ROUND_PATTERN.match(h):
            val = ws.cell(target_row, col_idx + 1).value
            history[h] = str(val or "").strip()

    wb.close()
    return history


# ──────────────────────────────────────────────
# 内部工具
# ──────────────────────────────────────────────

def _find_last_round_col(headers: list[str]) -> Optional[int]:
    """找到最后一个"第X轮勘测结果"列的位置（0-based）。"""
    last = None
    for idx, h in enumerate(headers):
        if ROUND_PATTERN.match(h):
            last = idx
    return last + 1 if last is not None else None  # 转为 1-based
