"""
gkclaw 邮件链路 · 离线契约回归（EVAL-STANDARDS 防假绿：纯本地、无网络、无 LLM）

覆盖：ids / schema / package（建包·解析·防篡改·防路径逃逸）/ registry 状态机与幂等 /
mapping 字段映射与分簇 / dispatch / ingest（ack·staged·final·error·冲突·隔离·双源）。

跑：
    agent\\.venv\\Scripts\\python.exe agent\\evals\\eval_gkclaw.py
    # --fixture 兼容 CI 入参（本评测恒离线，行为相同）
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

PROJECT_ROOT = Path(__file__).resolve().parents[2]   # aida/
sys.path.insert(0, str(PROJECT_ROOT))

TESTS = []


def test(fn):
    TESTS.append(fn)
    return fn


def tmpdir() -> Path:
    return Path(tempfile.mkdtemp(prefix="gkclaw-eval-"))


# ─── ids ───

@test
def ids_task_id_format_and_uniqueness():
    from agent.skills.zhgk.services.gkclaw import ids
    root = tmpdir()
    t1 = ids.new_task_id("K1903", root)
    t2 = ids.new_task_id("K1903", root)
    assert ids.is_valid_task_id(t1), f"task_id 不合契约正则: {t1}"
    assert t1 != t2, "连续生成必须唯一"
    assert t1.startswith("task-") and "K1903" in t1, t1
    assert t1 < t2, "序列应单调递增"


@test
def ids_sanitize_project_code():
    from agent.skills.zhgk.services.gkclaw import ids
    assert ids.sanitize_code("K1903") == "K1903"
    assert ids.sanitize_code("智算K19/03") == "K1903"
    assert ids.sanitize_code("！@#") == "ZHGK", "全非法字符回退 ZHGK"
    assert ids.sanitize_code("") == "ZHGK"


@test
def ids_package_id_format():
    from agent.skills.zhgk.services.gkclaw import ids
    p1, p2 = ids.new_package_id(), ids.new_package_id()
    assert p1.startswith("pkg-") and p1 != p2


@test
def ids_zip_path_safety():
    from agent.skills.zhgk.services.gkclaw import ids
    assert ids.is_safe_zip_path("assets/items/1/a.jpg")
    assert ids.is_safe_zip_path("task.json")
    assert not ids.is_safe_zip_path("../escape.txt")
    assert not ids.is_safe_zip_path("a/../../b")
    assert not ids.is_safe_zip_path("/abs/path")
    assert not ids.is_safe_zip_path("C:\\win\\path")
    assert not ids.is_safe_zip_path("")


# ─── schema ───

def _valid_task() -> dict:
    return {
        "task_id": "task-20260611-K1903-000001",
        "task_name": "A 机房 现场勘测",
        "project": {"project_id": "ACT001", "project_code": "K1903",
                    "project_name": "智算 Q3 · 客户甲一期"},
        "assignees": [{"surveyor_name": "张三", "surveyor_code": "S001"}],
        "items": [
            {"问题序号": "1", "勘测项": "机房门口标识是否清晰", "选项列表": [],
             "勘测结果": "", "to_front_备注": "", "to_back_备注": "", "示例图": [],
             "metadata": {"细分场景": "硬装入场"}},
            {"问题序号": "2", "勘测项": "接地线规格检查", "选项列表": [],
             "勘测结果": "", "to_front_备注": "", "to_back_备注": "", "示例图": [],
             "metadata": {}},
        ],
        "item_clusters": [
            {"cluster_id": "cluster-all", "cluster_name": "A 机房", "item_keys": ["1", "2"]},
        ],
        "dependency_rules": [],
        "supplemental_context": None,
        "metadata": {},
    }


@test
def schema_valid_task_passes():
    from agent.skills.zhgk.services.gkclaw import schema
    assert schema.validate_task(_valid_task()) == []


@test
def schema_task_required_fields():
    from agent.skills.zhgk.services.gkclaw import schema
    t = _valid_task(); t["task_id"] = "!bad id"
    assert any("task_id" in e for e in schema.validate_task(t))
    t = _valid_task(); t["assignees"] = []
    assert any("assignees" in e for e in schema.validate_task(t))
    t = _valid_task(); t["project"]["project_name"] = ""
    assert any("project_name" in e for e in schema.validate_task(t))
    t = _valid_task(); t["items"] = []
    assert any("items" in e for e in schema.validate_task(t))
    t = _valid_task(); t["item_clusters"] = []
    assert any("item_clusters" in e for e in schema.validate_task(t))


@test
def schema_task_item_and_cluster_rules():
    from agent.skills.zhgk.services.gkclaw import schema
    # 问题序号重复
    t = _valid_task(); t["items"][1]["问题序号"] = "1"
    assert any("问题序号" in e and "唯一" in e for e in schema.validate_task(t))
    # cluster 引用不存在的 key
    t = _valid_task(); t["item_clusters"][0]["item_keys"] = ["1", "99"]
    assert any("99" in e for e in schema.validate_task(t))
    # 有 item 不在任何簇
    t = _valid_task(); t["item_clusters"][0]["item_keys"] = ["1"]
    assert any("2" in e and "簇" in e for e in schema.validate_task(t))
    # cluster_id 重复
    t = _valid_task()
    t["item_clusters"] = [
        {"cluster_id": "c1", "cluster_name": "x", "item_keys": ["1"]},
        {"cluster_id": "c1", "cluster_name": "y", "item_keys": ["2"]},
    ]
    assert any("cluster_id" in e for e in schema.validate_task(t))
    # 示例图路径逃逸
    t = _valid_task()
    t["items"][0]["示例图"] = [{"asset_id": "a1", "path": "../evil.jpg", "mime_type": "image/jpeg"}]
    assert any("path" in e for e in schema.validate_task(t))


@test
def schema_task_dependency_rules():
    from agent.skills.zhgk.services.gkclaw import schema
    t = _valid_task()
    t["dependency_rules"] = [{
        "rule_id": "r1", "description": "d", "trigger_item_keys": ["1"],
        "trigger_semantics": "第1项表达不需要", "target_item_keys": ["2"],
        "action": {"type": "mark_not_applicable", "result": "不涉及"},
    }]
    assert schema.validate_task(t) == []
    t["dependency_rules"][0]["target_item_keys"] = ["88"]
    assert any("88" in e for e in schema.validate_task(t))
    t = _valid_task()
    t["dependency_rules"] = [{
        "rule_id": "r1", "trigger_item_keys": ["1"], "trigger_semantics": "x",
        "target_item_keys": ["2"], "action": {"type": "delete_item"},
    }]
    assert any("action" in e for e in schema.validate_task(t))


@test
def schema_manifest_ack_result_error():
    from agent.skills.zhgk.services.gkclaw import schema
    m = {"schema_version": "gkclaw.mail.v1", "package_id": "pkg-1", "package_type": "task.dispatch",
         "created_at": "2026-06-11T03:00:00+00:00", "source": "back-agent", "target": "front-agent",
         "task_id": "task-20260611-K1903-000001", "project_id": "ACT001", "project_code": "K1903",
         "checksum": {"task.json": "sha256:abc"}}
    assert schema.validate_manifest(m) == []
    bad = dict(m); bad["schema_version"] = "v999"
    assert any("schema_version" in e for e in schema.validate_manifest(bad))
    bad = dict(m); bad["package_type"] = "task.cancel"
    assert any("package_type" in e for e in schema.validate_manifest(bad))
    bad = dict(m); bad["checksum"] = {"../x": "sha256:a"}
    assert any("checksum" in e for e in schema.validate_manifest(bad))

    ack = {"status": "accepted", "task_id": "t1", "web_access_url": "https://f/x"}
    assert schema.validate_ack(ack) == []
    assert any("web_access_url" in e for e in schema.validate_ack({"status": "accepted", "task_id": "t1"}))

    res = {"task_id": "t1", "session": {"status": "completed"},
           "submitted_by": {"surveyor_name": "张三", "surveyor_code": "S001"}, "items": []}
    assert schema.validate_result(res) == []
    assert any("session" in e for e in schema.validate_result({"task_id": "t1", "items": []}))
    assert any("submitted_by" in e for e in schema.validate_result(
        {"task_id": "t1", "session": {"status": "active"}, "items": []}))

    err = {"task_id": "t1", "code": "invalid_task_payload", "message": "x"}
    assert schema.validate_error(err) == []
    assert any("code" in e for e in schema.validate_error({"task_id": "t1", "message": "x"}))


# ─── package ───

@test
def package_build_parse_roundtrip():
    from agent.skills.zhgk.services.gkclaw import package
    out = tmpdir()
    task = _valid_task()
    zp = package.build_package(
        package_type="task.dispatch", task_id=task["task_id"],
        project=task["project"], payload=task,
        assets={"assets/items/1/example-1.jpg": b"\xff\xd8fakejpg"},
        out_dir=out,
    )
    assert zp.name == f"task-{task['task_id']}.zip"
    parsed = package.parse_package(zp)
    assert parsed["ok"], parsed["errors"]
    assert parsed["manifest"]["package_type"] == "task.dispatch"
    assert parsed["payload"]["task_id"] == task["task_id"]
    assert parsed["payload_name"] == "task.json"
    assert "assets/items/1/example-1.jpg" in parsed["manifest"]["checksum"]


@test
def package_checksum_tamper_detected():
    import zipfile
    from agent.skills.zhgk.services.gkclaw import package
    out = tmpdir()
    task = _valid_task()
    zp = package.build_package(package_type="task.dispatch", task_id=task["task_id"],
                               project=task["project"], payload=task, out_dir=out)
    # 重写 task.json 内容但保留 manifest → checksum 不匹配
    tampered = out / "tampered.zip"
    with zipfile.ZipFile(zp) as src, zipfile.ZipFile(tampered, "w") as dst:
        for name in src.namelist():
            data = src.read(name)
            if name == "task.json":
                data = data.replace(b"K1903", b"HACKED")
            dst.writestr(name, data)
    parsed = package.parse_package(tampered)
    assert not parsed["ok"]
    assert any("checksum" in e for e in parsed["errors"])


@test
def package_path_escape_rejected():
    import json, zipfile
    from agent.skills.zhgk.services.gkclaw import package
    out = tmpdir()
    evil = out / "evil.zip"
    manifest = {"schema_version": "gkclaw.mail.v1", "package_id": "pkg-x",
                "package_type": "task.result", "created_at": "2026-06-11T00:00:00+00:00",
                "source": "front-agent", "target": "back-agent", "task_id": "t1",
                "checksum": {"result.json": "sha256:dead"}}
    with zipfile.ZipFile(evil, "w") as zf:
        zf.writestr("manifest.json", json.dumps(manifest))
        zf.writestr("result.json", "{}")
        zf.writestr("../escape.txt", "evil")
    parsed = package.parse_package(evil)
    assert not parsed["ok"]
    assert any("路径" in e or "path" in e.lower() for e in parsed["errors"])


@test
def package_missing_manifest_and_type_mismatch():
    import json, zipfile, hashlib
    from agent.skills.zhgk.services.gkclaw import package
    out = tmpdir()
    z1 = out / "nomanifest.zip"
    with zipfile.ZipFile(z1, "w") as zf:
        zf.writestr("task.json", "{}")
    parsed = package.parse_package(z1)
    assert not parsed["ok"] and any("manifest" in e for e in parsed["errors"])

    # 类型与 payload 文件不匹配：声明 task.result 但只有 task.json
    z2 = out / "mismatch.zip"
    payload = json.dumps({"x": 1})
    digest = "sha256:" + hashlib.sha256(payload.encode()).hexdigest()
    manifest = {"schema_version": "gkclaw.mail.v1", "package_id": "pkg-y",
                "package_type": "task.result", "created_at": "2026-06-11T00:00:00+00:00",
                "source": "front-agent", "target": "back-agent", "task_id": "t1",
                "checksum": {"task.json": digest}}
    with zipfile.ZipFile(z2, "w") as zf:
        zf.writestr("manifest.json", json.dumps(manifest))
        zf.writestr("task.json", payload)
    parsed = package.parse_package(z2)
    assert not parsed["ok"] and any("result.json" in e for e in parsed["errors"])


@test
def package_unlisted_files_policy():
    import json, zipfile, hashlib
    from agent.skills.zhgk.services.gkclaw import package
    out = tmpdir()
    payload = json.dumps({"task_id": "t1", "session": {"status": "active"},
                          "submitted_by": {"surveyor_code": "S001"}, "items": []})
    digest = "sha256:" + hashlib.sha256(payload.encode()).hexdigest()
    manifest = {"schema_version": "gkclaw.mail.v1", "package_id": "pkg-z",
                "package_type": "task.result", "created_at": "2026-06-11T00:00:00+00:00",
                "source": "front-agent", "target": "back-agent", "task_id": "t1",
                "checksum": {"result.json": digest}}
    # evidence 未列入 manifest → 仅 warning；根目录未列文件 → error
    z1 = out / "evidence-unlisted.zip"
    with zipfile.ZipFile(z1, "w") as zf:
        zf.writestr("manifest.json", json.dumps(manifest))
        zf.writestr("result.json", payload)
        zf.writestr("evidence/photo1.jpg", b"img")
    parsed = package.parse_package(z1)
    assert parsed["ok"], parsed["errors"]
    assert any("evidence" in w for w in parsed["warnings"])

    z2 = out / "root-unlisted.zip"
    with zipfile.ZipFile(z2, "w") as zf:
        zf.writestr("manifest.json", json.dumps(manifest))
        zf.writestr("result.json", payload)
        zf.writestr("extra.json", "{}")
    parsed = package.parse_package(z2)
    assert not parsed["ok"] and any("extra.json" in e for e in parsed["errors"])


@test
def package_extract_evidence():
    from agent.skills.zhgk.services.gkclaw import package
    out = tmpdir()
    res = {"task_id": "task-20260611-K1903-000009", "session": {"status": "completed"},
           "submitted_by": {"surveyor_name": "张三", "surveyor_code": "S001"}, "items": []}
    zp = package.build_package(package_type="task.result", task_id=res["task_id"],
                               project={"project_id": "ACT001", "project_code": "K1903"},
                               payload=res, assets={"evidence/p1.jpg": b"img1"}, out_dir=out)
    dest = out / "ev"
    saved = package.extract_files(zp, ["evidence/p1.jpg"], dest)
    assert saved[0].read_bytes() == b"img1"
    assert saved[0] == dest / "evidence" / "p1.jpg"


# ─── registry ───

def _registry(root=None):
    from agent.skills.zhgk.services.gkclaw.registry import TaskRegistry
    return TaskRegistry(root or tmpdir())


def _create_task(reg, tid="task-20260611-K1903-000001"):
    t = _valid_task()
    t["task_id"] = tid
    return reg.create_task(
        task_id=tid, task_payload=t, zip_path="outbox/x.zip",
        table_fingerprint="sha256:tbl", project=t["project"], dry_run=False,
    )


@test
def registry_create_get_and_states():
    reg = _registry()
    task = _create_task(reg)
    assert task["state"] == "planned"
    reg.set_state(task["task_id"], "dispatched", mailgw_task_id="m1")
    got = reg.get(task["task_id"])
    assert got["state"] == "dispatched" and got["mailgw_task_id"] == "m1"
    assert reg.get("task-not-exists") is None
    assert [t["task_id"] for t in reg.list_tasks()] == [task["task_id"]]
    reg.mark_superseded(task["task_id"])
    assert reg.get(task["task_id"])["state"] == "superseded"


@test
def registry_decide_inbound_idempotency_matrix():
    """契约 §20 幂等/冲突矩阵 + 边界规则。"""
    from agent.skills.zhgk.services.gkclaw.registry import decide_inbound
    base_task = {"task_id": "t1", "state": "dispatched"}
    known = [{"package_id": "p1", "checksum": "sha256:aaa",
              "package_type": "task.import_ack", "disposition": "processed"}]

    # 未知 task_id → quarantine
    d, _ = decide_inbound(None, [], package_type="task.import_ack",
                          package_id="px", checksum="sha256:x", session_status="")
    assert d == "quarantine"
    # 同 package_id 同 checksum → duplicate
    d, _ = decide_inbound(base_task, known, package_type="task.import_ack",
                          package_id="p1", checksum="sha256:aaa", session_status="")
    assert d == "duplicate"
    # 同 package_id 异 checksum → conflict
    d, _ = decide_inbound(base_task, known, package_type="task.import_ack",
                          package_id="p1", checksum="sha256:bbb", session_status="")
    assert d == "conflict"
    # superseded 任务 → archived（落档不生效）
    d, _ = decide_inbound({"task_id": "t1", "state": "superseded"}, [],
                          package_type="task.result", package_id="p2",
                          checksum="sha256:c", session_status="completed")
    assert d == "archived"
    # dispatched 收 ack → processed
    d, _ = decide_inbound(base_task, [], package_type="task.import_ack",
                          package_id="p2", checksum="sha256:c", session_status="")
    assert d == "processed"
    # accepted 收 staged result → processed
    d, _ = decide_inbound({"task_id": "t1", "state": "accepted"}, [],
                          package_type="task.result", package_id="p3",
                          checksum="sha256:d", session_status="active")
    assert d == "processed"
    # completed 后收 staged → quarantine（final 后 staged）
    d, _ = decide_inbound({"task_id": "t1", "state": "completed"}, [],
                          package_type="task.result", package_id="p4",
                          checksum="sha256:e", session_status="active")
    assert d == "quarantine"
    # completed 后收异内容 final → conflict
    d, _ = decide_inbound({"task_id": "t1", "state": "completed"}, [],
                          package_type="task.result", package_id="p5",
                          checksum="sha256:f", session_status="completed")
    assert d == "conflict"
    # error 包任意状态 → processed
    d, _ = decide_inbound(base_task, [], package_type="task.error",
                          package_id="p6", checksum="sha256:g", session_status="")
    assert d == "processed"


@test
def registry_result_versions_and_notes():
    reg = _registry()
    task = _create_task(reg, "task-20260611-K1903-000002")
    p1 = reg.store_result_version(task["task_id"], {"session": {"status": "active"}, "items": []},
                                  final=False)
    p2 = reg.store_result_version(task["task_id"], {"session": {"status": "completed"}, "items": []},
                                  final=True)
    assert p1.name == "result-001.json" and p2.name == "result-002.json"
    got = reg.get(task["task_id"])
    assert got["result_versions"] == 2 and got["final_result"] == "results/result-002.json"
    reg.append_result_notes(task["task_id"], [{"问题序号": "5", "to_back_备注": "[规则自动处理] x"}])
    notes = reg.read_result_notes(task["task_id"])
    assert notes and notes[0]["问题序号"] == "5"


@test
def registry_package_ledger_and_scan_ledger():
    reg = _registry()
    task = _create_task(reg, "task-20260611-K1903-000003")
    reg.record_package(task["task_id"], {"package_id": "p1", "checksum": "sha256:a",
                                         "package_type": "task.import_ack",
                                         "direction": "in", "disposition": "processed",
                                         "mail_id": 7})
    pkgs = reg.packages(task["task_id"])
    assert len(pkgs) == 1 and pkgs[0]["package_id"] == "p1" and pkgs[0]["at"]
    # 扫描账本
    assert reg.scanned_mail_ids() == set()
    reg.mark_mail_scanned(7, "processed")
    reg.mark_mail_scanned(8, "ignored")
    assert reg.scanned_mail_ids() == {7, 8}


# ─── mapping ───

def _survey_rows() -> list[dict]:
    """模拟 read_survey_table 输出（列结构见 survey_table_builder.RESULT_TABLE_HEADERS）。"""
    return [
        {"序号": 1, "细分场景": "硬装入场", "勘测要素": "接地", "项目": "接地线",
         "检查内容": "检查机房接地线规格是否达标", "勘测方法": "现场勘测",
         "最新检查结果": "", "AI评估结果": "", "备注": ""},
        {"序号": 2, "细分场景": "硬装入场", "勘测要素": "层高", "项目": "梁下净高",
         "检查内容": "测量梁下净高是否≥3m", "勘测方法": "数据",
         "最新检查结果": "", "AI评估结果": "", "备注": ""},
        {"序号": 3, "细分场景": "通液前", "勘测要素": "管路", "项目": "排水管",
         "检查内容": "检查排水管走向与坡度", "勘测方法": "现场勘测",
         "最新检查结果": "", "AI评估结果": "", "备注": ""},
    ]


def _base_items() -> list[dict]:
    """模拟 load_base_table 输出（含建表时被丢弃的背景知识列）。"""
    return [
        {"序号": 11, "代际_制冷": "G5-液冷", "分类": "标准", "细分场景": "硬装入场",
         "勘测要素": "接地", "项目": "接地线", "检查内容": "检查机房接地线规格是否达标",
         "是否支持视频勘测": "是", "勘测方法": "现场勘测", "检查结果": "", "备注": "",
         "语音助手背景知识": "接地线常见 35/50mm²", "视频勘测背景知识": "拍摄接地排与标识牌"},
        {"序号": 12, "代际_制冷": "G5-液冷", "分类": "标准", "细分场景": "通液前",
         "勘测要素": "管路", "项目": "排水管", "检查内容": "检查排水管走向与坡度",
         "是否支持视频勘测": "否", "勘测方法": "现场勘测", "检查结果": "", "备注": "",
         "语音助手背景知识": "", "视频勘测背景知识": "沿管路全程拍摄"},
    ]


def _project() -> dict:
    return {"project_code": "K1903", "project_name": "智算 Q3 · 客户甲一期",
            "room_name": "A 机房", "activity_id": "ACT001"}


@test
def mapping_builds_contract_task():
    from agent.skills.zhgk.services.gkclaw import mapping, schema
    task = mapping.build_task_payload(
        task_id="task-20260611-K1903-000001",
        rows=_survey_rows(), base_items=_base_items(),
        project=_project(),
        assignees=[{"surveyor_name": "张三", "surveyor_code": "S001"}],
    )
    assert schema.validate_task(task) == [], schema.validate_task(task)
    # 只下发 现场勘测 条目（序号 2 是数据类，不下发）
    keys = [it["问题序号"] for it in task["items"]]
    assert keys == ["1", "3"]
    it1 = task["items"][0]
    assert it1["勘测项"] == "检查机房接地线规格是否达标"
    assert it1["选项列表"] == [] and it1["勘测结果"] == "" and it1["to_back_备注"] == ""
    # to_front_备注 = 【勘测要素/项目】+ 视频勘测背景知识 + 语音助手背景知识
    assert it1["to_front_备注"].startswith("【接地/接地线】")
    assert "拍摄接地排与标识牌" in it1["to_front_备注"]
    assert "接地线常见 35/50mm²" in it1["to_front_备注"]
    # metadata 溯源
    assert it1["metadata"]["细分场景"] == "硬装入场"
    assert it1["metadata"]["是否支持视频勘测"] == "是"
    assert it1["metadata"]["底表序号"] == 11
    # 顶层
    assert task["task_name"] == "智算 Q3 · 客户甲一期·A 机房 现场勘测"
    assert task["project"] == {"project_id": "ACT001", "project_code": "K1903",
                               "project_name": "智算 Q3 · 客户甲一期"}
    assert task["dependency_rules"] == [] and task["supplemental_context"] is None
    assert task["metadata"]["generation_cooling"] == "" and task["metadata"]["room_name"] == "A 机房"


@test
def mapping_join_miss_is_tolerated():
    from agent.skills.zhgk.services.gkclaw import mapping
    task = mapping.build_task_payload(
        task_id="task-20260611-K1903-000002",
        rows=_survey_rows(), base_items=[],  # 底表 join 全失败
        project=_project(),
        assignees=[{"surveyor_name": "张三", "surveyor_code": "S001"}],
    )
    it1 = task["items"][0]
    assert it1["to_front_备注"] == "【接地/接地线】"
    assert "底表序号" not in it1["metadata"]


@test
def mapping_clusters_default_single():
    """物理空间分簇：v1 无字段 → 单一兜底簇 cluster-all（cluster_name=机房名）。"""
    from agent.skills.zhgk.services.gkclaw import mapping
    clusters = mapping.derive_clusters(["1", "3"], room_name="A 机房")
    assert clusters == [{"cluster_id": "cluster-all", "cluster_name": "A 机房",
                         "item_keys": ["1", "3"]}]
    assert mapping.derive_clusters(["1"], room_name="")[0]["cluster_name"] == "全部条目"


@test
def mapping_clusters_by_field_when_present():
    """预留升级点：提供 cluster_values（key→物理位置标签）时按标签分簇 + 空值落兜底。"""
    from agent.skills.zhgk.services.gkclaw import mapping
    clusters = mapping.derive_clusters(
        ["1", "2", "3", "4"], room_name="A 机房",
        cluster_values={"1": "配电区", "2": "制冷区", "3": "配电区", "4": ""},
    )
    assert clusters == [
        {"cluster_id": "cluster-01", "cluster_name": "配电区", "item_keys": ["1", "3"]},
        {"cluster_id": "cluster-02", "cluster_name": "制冷区", "item_keys": ["2"]},
        {"cluster_id": "cluster-other", "cluster_name": "其他", "item_keys": ["4"]},
    ]


# ─── mailer mailgw backend ───

class _FakeHttpResp:
    def __init__(self, data: dict):
        import json as _json
        self._raw = _json.dumps(data).encode()
    def read(self):
        return self._raw
    def __enter__(self):
        return self
    def __exit__(self, *a):
        return False


@test
def mailer_mailgw_backend_sends():
    import os, json as _json
    import agent.mailer as mailer
    os.environ["MAILGW_BASE"] = "http://127.0.0.1:8025"
    os.environ["MAILGW_TOKEN"] = "tok-aida"
    os.environ["AIDA_MAIL_BACKEND"] = "mailgw"
    captured: dict = {}

    def fake_urlopen(req, timeout=0):
        captured["url"] = req.full_url
        captured["auth"] = req.get_header("Authorization")
        captured["body"] = _json.loads(req.data.decode())
        return _FakeHttpResp({"task_id": "mgw-1", "status": "pending_approval",
                              "message": "已转入待审批队列"})

    orig = mailer.urllib.request.urlopen
    mailer.urllib.request.urlopen = fake_urlopen
    try:
        r = mailer.send_mail(["front@corp.com"], "[GKCLAW][TASK_DISPATCH] K1903/t1",
                             "正文", attachments=["D:/x.zip"], dry_run=False)
    finally:
        mailer.urllib.request.urlopen = orig
        os.environ.pop("AIDA_MAIL_BACKEND", None)
    assert r["ok"] and r["via"] == "mailgw"
    assert r["mailgw_task_id"] == "mgw-1" and r["mailgw_status"] == "pending_approval"
    assert captured["url"].endswith("/api/send")
    assert captured["auth"] == "Bearer tok-aida"
    assert captured["body"]["to"] == ["front@corp.com"]
    assert captured["body"]["attachments"] == ["D:/x.zip"]


@test
def mailer_mailgw_requires_token():
    import os
    import agent.mailer as mailer
    os.environ.pop("MAILGW_TOKEN", None)
    os.environ["AIDA_MAIL_BACKEND"] = "mailgw"
    try:
        r = mailer.send_mail("a@b.com", "s", "b", dry_run=False)
    finally:
        os.environ.pop("AIDA_MAIL_BACKEND", None)
    assert not r["ok"] and "MAILGW_TOKEN" in r["error"]


# ─── mailbox ───

@test
def mailbox_list_inbox_request_shape():
    import os
    import agent.mailbox as mailbox
    os.environ["MAILGW_BASE"] = "http://127.0.0.1:8025"
    os.environ["MAILGW_TOKEN"] = "tok-aida"
    assert mailbox.is_configured()
    captured: dict = {}

    def fake_urlopen(req, timeout=0):
        captured["url"] = req.full_url
        captured["auth"] = req.get_header("Authorization")
        captured["method"] = req.get_method()
        return _FakeHttpResp({"new_count": 1, "mails": [
            {"mail_id": 3, "from": "front@corp.com",
             "subject": "[GKCLAW][TASK_IMPORT_ACK] K1903/t1",
             "date": "Thu, 11 Jun 2026 09:00:00 +0800",
             "snippet": "见附件", "has_attachments": True, "is_read": False}]})

    orig = mailbox.urllib.request.urlopen
    mailbox.urllib.request.urlopen = fake_urlopen
    try:
        data = mailbox.list_inbox(refresh=True, limit=50, unread_only=False)
    finally:
        mailbox.urllib.request.urlopen = orig
    assert data["mails"][0]["mail_id"] == 3
    assert captured["method"] == "GET" and captured["auth"] == "Bearer tok-aida"
    assert "/api/inbox" in captured["url"]
    assert "refresh=true" in captured["url"] and "limit=50" in captured["url"]


@test
def mailbox_save_attachment():
    import os, json as _json
    import agent.mailbox as mailbox
    os.environ["MAILGW_BASE"] = "http://127.0.0.1:8025"
    os.environ["MAILGW_TOKEN"] = "tok-aida"
    captured: dict = {}

    def fake_urlopen(req, timeout=0):
        captured["url"] = req.full_url
        captured["body"] = _json.loads(req.data.decode())
        return _FakeHttpResp({"saved_to": "D:/tmp/ack-t1.zip"})

    orig = mailbox.urllib.request.urlopen
    mailbox.urllib.request.urlopen = fake_urlopen
    try:
        saved = mailbox.save_attachment(3, 0, "D:/tmp")
    finally:
        mailbox.urllib.request.urlopen = orig
    assert saved == "D:/tmp/ack-t1.zip"
    assert captured["url"].endswith("/api/inbox/3/attachments/0/save")
    assert captured["body"] == {"save_path": "D:/tmp"}


# ─── dispatch ───

def _make_survey_xlsx(out_dir) -> str:
    """用真实 builder 生成全量勘测结果表（3 条：2 现场 + 1 数据）。"""
    from agent.skills.zhgk.services.survey_table_builder import build_survey_table
    items = [
        {"细分场景": "硬装入场", "勘测要素": "接地", "项目": "接地线",
         "检查内容": "检查机房接地线规格是否达标", "勘测方法": "现场勘测", "备注": ""},
        {"细分场景": "硬装入场", "勘测要素": "层高", "项目": "梁下净高",
         "检查内容": "测量梁下净高是否≥3m", "勘测方法": "数据", "备注": ""},
        {"细分场景": "通液前", "勘测要素": "管路", "项目": "排水管",
         "检查内容": "检查排水管走向与坡度", "勘测方法": "现场勘测", "备注": ""},
    ]
    return build_survey_table(items, str(out_dir), "ACT001", "智算 Q3 · 客户甲一期", "A 机房")


_ASSIGNEES = [{"surveyor_name": "张三", "surveyor_code": "S001"}]


@test
def dispatch_dry_run_creates_task_and_zip():
    from agent.skills.zhgk.services.gkclaw import dispatch, package
    from agent.skills.zhgk.services.gkclaw.registry import TaskRegistry
    root = tmpdir()
    table = _make_survey_xlsx(root / "Output")
    r = dispatch.dispatch_task(
        runtime_dir=root / "RunTime", survey_table_path=table,
        project=_project(), assignees=_ASSIGNEES, dry_run=True,
    )
    assert r["dry_run"] is True and r["state"] == "dispatched"
    assert r["items_count"] == 2
    reg = TaskRegistry(root / "RunTime")
    task = reg.get(r["task_id"])
    assert task["state"] == "dispatched" and task["dry_run"] is True
    assert task["table_fingerprint"].startswith("sha256:")
    zp = reg.root / r["task_id"] / "outbox" / f"task-{r['task_id']}.zip"
    assert zp.exists()
    parsed = package.parse_package(zp)
    assert parsed["ok"] and parsed["payload"]["task_id"] == r["task_id"]
    assert len(reg.packages(r["task_id"])) == 1  # 出站包入账本


@test
def dispatch_real_send_and_supersede():
    from agent.skills.zhgk.services.gkclaw import dispatch
    from agent.skills.zhgk.services.gkclaw.registry import TaskRegistry
    root = tmpdir()
    table = _make_survey_xlsx(root / "Output")
    sent: list[dict] = []

    def fake_send(to, subject, body, *, attachments=None, dry_run=None):
        sent.append({"to": to, "subject": subject, "attachments": attachments})
        return {"ok": True, "via": "mailgw", "mailgw_task_id": "mgw-7", "mailgw_status": "sent"}

    r1 = dispatch.dispatch_task(
        runtime_dir=root / "RunTime", survey_table_path=table,
        project=_project(), assignees=_ASSIGNEES, dry_run=True,
    )
    r2 = dispatch.dispatch_task(
        runtime_dir=root / "RunTime", survey_table_path=table,
        project=_project(), assignees=_ASSIGNEES,
        dry_run=False, send_fn=fake_send,
        frontagent_mailbox="front@corp.com", previous_task_id=r1["task_id"],
    )
    reg = TaskRegistry(root / "RunTime")
    assert reg.get(r1["task_id"])["state"] == "superseded"          # 重发=新 task_id+旧任务取代
    task2 = reg.get(r2["task_id"])
    assert task2["state"] == "dispatched" and task2["mailgw_task_id"] == "mgw-7"
    assert sent[0]["to"] == ["front@corp.com"]
    assert sent[0]["subject"] == f"[GKCLAW][TASK_DISPATCH] K1903/{r2['task_id']}"
    assert sent[0]["attachments"][0].endswith(f"task-{r2['task_id']}.zip")


@test
def dispatch_send_failure_marks_failed():
    from agent.skills.zhgk.services.gkclaw import dispatch
    from agent.skills.zhgk.services.gkclaw.registry import TaskRegistry
    root = tmpdir()
    table = _make_survey_xlsx(root / "Output")

    def bad_send(to, subject, body, *, attachments=None, dry_run=None):
        return {"ok": False, "error": "connection refused", "via": "mailgw"}

    try:
        dispatch.dispatch_task(
            runtime_dir=root / "RunTime", survey_table_path=table,
            project=_project(), assignees=_ASSIGNEES,
            dry_run=False, send_fn=bad_send, frontagent_mailbox="front@corp.com",
        )
        assert False, "应抛 RuntimeError"
    except RuntimeError as e:
        assert "connection refused" in str(e)
    reg = TaskRegistry(root / "RunTime")
    tasks = reg.list_tasks()
    assert len(tasks) == 1 and tasks[0]["state"] == "failed"
    assert "connection refused" in tasks[0]["last_error"]


# ─── ingest ───

def _dispatched_env():
    """构造已下发环境：返回 (root, reg, task_id, task_payload, survey_table_path, input_dir)。"""
    from agent.skills.zhgk.services.gkclaw import dispatch
    from agent.skills.zhgk.services.gkclaw.registry import TaskRegistry
    root = tmpdir()
    (root / "Input").mkdir()
    table = _make_survey_xlsx(root / "Output")
    r = dispatch.dispatch_task(runtime_dir=root / "RunTime", survey_table_path=table,
                               project=_project(), assignees=_ASSIGNEES, dry_run=True)
    reg = TaskRegistry(root / "RunTime")
    return root, reg, r["task_id"], reg.task_payload(r["task_id"]), table, root / "Input"


def _ack_zip(task_payload, out, *, assignees=None, package_id=None):
    from agent.skills.zhgk.services.gkclaw import package
    ack = {"status": "accepted", "task_id": task_payload["task_id"],
           "task_name": task_payload["task_name"], "project": task_payload["project"],
           "assignees": assignees if assignees is not None else task_payload["assignees"],
           "web_access_url": "https://front-agent.example.com/tasks/web/wa_7b4f",
           "accepted_at": "2026-06-11T03:01:00.000Z"}
    return package.build_package(package_type="task.import_ack", task_id=task_payload["task_id"],
                                 project=task_payload["project"], payload=ack, out_dir=out,
                                 source="front-agent", target="back-agent", package_id=package_id)


def _result_zip(task_payload, out, *, status, items=None, submitted_code="S001",
                package_id=None, notes=False, evidence=False):
    from agent.skills.zhgk.services.gkclaw import package
    r_items = items if items is not None else [
        {"问题序号": "1", "勘测项": "检查机房接地线规格是否达标", "勘测结果": "满足",
         "to_back_备注": "[规则自动处理: r1] 语义等价不涉及" if notes else ""},
        {"问题序号": "3", "勘测项": "检查排水管走向与坡度", "勘测结果": "不满足"},
    ]
    res = {"task_id": task_payload["task_id"], "task_name": task_payload["task_name"],
           "project": task_payload["project"], "assignees": task_payload["assignees"],
           "submitted_by": {"surveyor_name": "张三", "surveyor_code": submitted_code},
           "session": {"task_id": task_payload["task_id"], "status": status},
           "items": r_items, "evidence": [], "updates": [], "observations": [],
           "exported_at": "2026-06-11T03:31:00.000Z"}
    assets = {"evidence/p1.jpg": b"img1"} if evidence else None
    return package.build_package(package_type="task.result", task_id=task_payload["task_id"],
                                 project=task_payload["project"], payload=res, out_dir=out,
                                 assets=assets, source="front-agent", target="back-agent",
                                 package_id=package_id)


@test
def ingest_ack_updates_state_and_mismatch_quarantines():
    from agent.skills.zhgk.services.gkclaw import ingest
    root, reg, tid, tp, table, input_dir = _dispatched_env()
    out = ingest.ingest_zip(_ack_zip(tp, tmpdir()), runtime_dir=root / "RunTime",
                            input_dir=input_dir, survey_table_path=table)
    assert out["disposition"] == "processed", out
    task = reg.get(tid)
    assert task["state"] == "accepted"
    assert task["web_access_url"].startswith("https://front-agent")
    # 人员不一致的 ACK → 隔离，状态不动
    bad = _ack_zip(tp, tmpdir(), assignees=[{"surveyor_name": "李四", "surveyor_code": "S099"}])
    out2 = ingest.ingest_zip(bad, runtime_dir=root / "RunTime",
                             input_dir=input_dir, survey_table_path=table)
    assert out2["disposition"] == "quarantine" and "不一致" in out2["note"]
    assert reg.get(tid)["state"] == "accepted"


@test
def ingest_staged_records_final_writes_filled_table():
    import openpyxl
    from agent.skills.zhgk.services.gkclaw import ingest
    root, reg, tid, tp, table, input_dir = _dispatched_env()
    ingest.ingest_zip(_ack_zip(tp, tmpdir()), runtime_dir=root / "RunTime",
                      input_dir=input_dir, survey_table_path=table)
    # staged：记录版本，不落 Input，不推进到 completed
    out = ingest.ingest_zip(_result_zip(tp, tmpdir(), status="active"),
                            runtime_dir=root / "RunTime", input_dir=input_dir,
                            survey_table_path=table)
    assert out["disposition"] == "processed"
    task = reg.get(tid)
    assert task["state"] == "staged_returned" and task["result_versions"] == 1
    assert not (input_dir / "已填写_全量勘测结果表.xlsx").exists()
    # final：转写已填写表落 Input + completed + 备注留档 + evidence 解出
    out = ingest.ingest_zip(_result_zip(tp, tmpdir(), status="completed",
                                        notes=True, evidence=True),
                            runtime_dir=root / "RunTime", input_dir=input_dir,
                            survey_table_path=table)
    assert out["disposition"] == "processed" and out["merged"] is True
    task = reg.get(tid)
    assert task["state"] == "completed" and task["final_result"]
    filled = input_dir / "已填写_全量勘测结果表.xlsx"
    assert filled.exists()
    ws = openpyxl.load_workbook(filled, data_only=True).active
    headers = [c.value for c in ws[1]]
    col = headers.index("最新检查结果") + 1
    assert ws.cell(2, col).value == "满足"      # 序号1
    assert ws.cell(4, col).value == "不满足"    # 序号3
    assert reg.read_result_notes(tid)[0]["问题序号"] == "1"
    assert (reg.root / tid / "evidence" / "p1.jpg").read_bytes() == b"img1"


@test
def ingest_idempotency_and_conflicts():
    from agent.skills.zhgk.services.gkclaw import ingest
    root, reg, tid, tp, table, input_dir = _dispatched_env()
    kw = dict(runtime_dir=root / "RunTime", input_dir=input_dir, survey_table_path=table)
    final1 = _result_zip(tp, tmpdir(), status="completed", package_id="pkg-final-1")
    assert ingest.ingest_zip(final1, **kw)["disposition"] == "processed"
    # 同一 ZIP 原包重放（同 package_id 同 checksum）→ duplicate（幂等成功）
    # 注意不能重新 build：created_at 变化会导致 zip checksum 不同 → 会判 conflict（符合契约）
    assert ingest.ingest_zip(final1, **kw)["disposition"] == "duplicate"
    # final 后异内容 final（新 package_id）→ conflict + 隔离落盘
    final2 = _result_zip(tp, tmpdir(), status="completed",
                         items=[{"问题序号": "1", "勘测结果": "不涉及"}])
    out = ingest.ingest_zip(final2, **kw)
    assert out["disposition"] == "conflict"
    assert any(reg.quarantine_dir().iterdir())
    # final 后 staged → quarantine
    staged = _result_zip(tp, tmpdir(), status="active")
    assert ingest.ingest_zip(staged, **kw)["disposition"] == "quarantine"
    assert reg.get(tid)["state"] == "completed"  # 状态不被污染


@test
def ingest_submitted_by_guard():
    from agent.skills.zhgk.services.gkclaw import ingest
    root, reg, tid, tp, table, input_dir = _dispatched_env()
    bad = _result_zip(tp, tmpdir(), status="completed", submitted_code="S999")
    out = ingest.ingest_zip(bad, runtime_dir=root / "RunTime",
                            input_dir=input_dir, survey_table_path=table)
    assert out["disposition"] == "quarantine" and "submitted_by" in out["note"]
    assert reg.get(tid)["state"] == "dispatched"


@test
def ingest_dual_source_and_fingerprint_guard():
    from agent.skills.zhgk.services.gkclaw import ingest
    # 双源冲突：Input 已有待合并表 → 邮件结果暂存 pending_results，不覆盖
    root, reg, tid, tp, table, input_dir = _dispatched_env()
    (input_dir / "已填写_全量勘测结果表.xlsx").write_bytes(b"manual upload")
    out = ingest.ingest_zip(_result_zip(tp, tmpdir(), status="completed"),
                            runtime_dir=root / "RunTime", input_dir=input_dir,
                            survey_table_path=table)
    assert out["disposition"] == "processed" and out["merged"] is False
    assert "先到先得" in out["note"]
    assert any((reg.root / tid / "pending_results").iterdir())
    assert (input_dir / "已填写_全量勘测结果表.xlsx").read_bytes() == b"manual upload"
    # 表指纹漂移：下发后表被改 → final 不合并
    root, reg, tid, tp, table, input_dir = _dispatched_env()
    import openpyxl
    wb = openpyxl.load_workbook(table); wb.active.cell(2, 5, "被改过的检查内容"); wb.save(table)
    out = ingest.ingest_zip(_result_zip(tp, tmpdir(), status="completed"),
                            runtime_dir=root / "RunTime", input_dir=input_dir,
                            survey_table_path=table)
    assert out["merged"] is False and "指纹" in out["note"]
    assert not (input_dir / "已填写_全量勘测结果表.xlsx").exists()
    assert reg.get(tid)["state"] == "completed"  # 任务仍完成，仅合并阻塞


@test
def ingest_unknown_task_and_broken_zip():
    from agent.skills.zhgk.services.gkclaw import ingest
    from agent.skills.zhgk.services.gkclaw.registry import TaskRegistry
    root = tmpdir(); (root / "Input").mkdir()
    reg = TaskRegistry(root / "RunTime")
    # 未知 task_id：合法包但本地无任务 → 隔离留档，不创建任务
    fake_tp = _valid_task()
    out = ingest.ingest_zip(_ack_zip(fake_tp, tmpdir()), runtime_dir=root / "RunTime",
                            input_dir=root / "Input", survey_table_path=None)
    assert out["disposition"] == "quarantine"
    assert reg.get(fake_tp["task_id"]) is None
    # 坏 ZIP → 隔离不抛
    bad = root / "bad.zip"; bad.write_bytes(b"\xff\xfe not a zip")
    out = ingest.ingest_zip(bad, runtime_dir=root / "RunTime",
                            input_dir=root / "Input", survey_table_path=None)
    assert out["disposition"] == "quarantine"


@test
def poll_and_ingest_with_fake_mailbox():
    import shutil
    from agent.skills.zhgk.services.gkclaw import ingest
    root, reg, tid, tp, table, input_dir = _dispatched_env()
    ack_zip = _ack_zip(tp, tmpdir())

    class FakeMailbox:
        @staticmethod
        def is_configured(): return True
        @staticmethod
        def list_inbox(refresh=True, limit=50, unread_only=False):
            return {"new_count": 2, "mails": [
                {"mail_id": 1, "subject": "[GKCLAW][TASK_IMPORT_ACK] K1903/x",
                 "has_attachments": True},
                {"mail_id": 2, "subject": "普通邮件", "has_attachments": False},
            ]}
        @staticmethod
        def save_attachment(mail_id, index, save_dir):
            if mail_id == 1 and index == 0:
                dest = Path(save_dir) / ack_zip.name
                Path(save_dir).mkdir(parents=True, exist_ok=True)
                shutil.copy(ack_zip, dest)
                return str(dest)
            raise ingest.MailboxExhausted("404")

    summary = ingest.poll_and_ingest(runtime_dir=root / "RunTime", input_dir=input_dir,
                                     survey_table_path=table, mailbox_mod=FakeMailbox)
    assert summary["checked"] == 2 and summary["processed"] == 1
    assert reg.get(tid)["state"] == "accepted"
    assert reg.scanned_mail_ids() == {1, 2}
    # 第二次轮询：账本生效，全部跳过
    summary2 = ingest.poll_and_ingest(runtime_dir=root / "RunTime", input_dir=input_dir,
                                      survey_table_path=table, mailbox_mod=FakeMailbox)
    assert summary2["checked"] == 0


# ─── task_dispatch step ───

def _step_ctx(intent="survey_work", extra_project=None):
    """构造 step 级测试环境：work_root + 全量表 + project_info.json。"""
    import json as _json
    from agent.skills.base import SkillContext
    root = tmpdir()
    project = {"intent": intent, "project_code": "K1903",
               "project_name": "智算 Q3 · 客户甲一期", "room_name": "A 机房",
               "activity_id": "ACT001"}
    project.update(extra_project or {})
    ctx = SkillContext(skill_id="zhgk", work_root=root, run_id="r1", project=project)
    ctx.ensure_dirs()
    table = _make_survey_xlsx(ctx.output_dir)
    (ctx.runtime_dir / "project_info.json").write_text(
        _json.dumps({"survey_table_path": table}, ensure_ascii=False), encoding="utf-8")
    return ctx


@test
def step_guard_hitl_and_skip():
    from agent.skills.zhgk.steps.task_dispatch import TaskDispatchStep
    from agent.skills.zhgk.steps._intent_guard import should_skip
    # 仅 survey_work：supplement 守卫直接跳过（用户拍板）
    assert should_skip("task_dispatch", {"intent": "supplement"})
    assert should_skip("task_dispatch", {"intent": "report_gen"})
    assert not should_skip("task_dispatch", {"intent": "survey_work"})
    step = TaskDispatchStep()
    # 无决策 → ChoiceCard HITL（下发/跳过）
    ctx = _step_ctx()
    check = step.check_inputs(ctx)
    assert not check["ok"]
    values = [o["value"] for o in check["need_inputs"][0]["options"]]
    assert values == ["dispatch", "skip"]
    # 决策=skip → 放行，run 记录跳过
    ctx = _step_ctx(extra_project={"dispatch_decision": "skip"})
    assert step.check_inputs(ctx)["ok"]
    result = step.run(ctx, {}, lambda m: None)
    assert result["metrics"]["gkclaw_skipped"] is True


@test
def step_dispatch_dry_run_idempotent():
    import json as _json, os
    from agent.skills.zhgk.steps.task_dispatch import TaskDispatchStep
    from agent.skills.zhgk.services.gkclaw.registry import TaskRegistry
    os.environ.pop("AIDA_SEND_EMAIL", None)   # 默认 dry-run
    step = TaskDispatchStep()
    ctx = _step_ctx(extra_project={"dispatch_decision": "dispatch", "assignees": _ASSIGNEES})
    assert step.check_inputs(ctx)["ok"]
    logs = []
    result = step.run(ctx, {}, logs.append)
    m = result["metrics"]
    assert m["gkclaw_task_id"].startswith("task-") and m["gkclaw_dry_run"] is True
    assert m["gkclaw_state"] == "dispatched" and m["gkclaw_items"] == 2
    info = _json.loads((ctx.runtime_dir / "project_info.json").read_text(encoding="utf-8"))
    assert info["gkclaw_task_id"] == m["gkclaw_task_id"]
    # 再跑（resume 重放）→ 不重复下发，仍报同一任务状态
    result2 = step.run(ctx, {}, logs.append)
    assert result2["metrics"]["gkclaw_task_id"] == m["gkclaw_task_id"]
    assert len(TaskRegistry(ctx.runtime_dir).list_tasks()) == 1


@test
def step_missing_assignees_blocks():
    from agent.skills.zhgk.steps.task_dispatch import TaskDispatchStep
    step = TaskDispatchStep()
    ctx = _step_ctx(extra_project={"dispatch_decision": "dispatch"})  # 无 assignees
    check = step.check_inputs(ctx)
    assert not check["ok"]
    assert any("assignees.json" in x for x in check["missing"])


# ─── main ───

def main() -> int:
    failed = 0
    for fn in TESTS:
        try:
            fn()
            sys.stdout.write(f"  ✓ {fn.__name__}\n")
        except AssertionError as e:
            failed += 1
            sys.stdout.write(f"  ❌ {fn.__name__}: {e}\n")
        except Exception as e:  # noqa: BLE001
            failed += 1
            sys.stdout.write(f"  ❌ {fn.__name__}: {type(e).__name__}: {e}\n")
    status = "OK" if failed == 0 else "FAIL"
    sys.stdout.write(f"[eval-gkclaw] {status} · {len(TESTS) - failed}/{len(TESTS)}\n")
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
