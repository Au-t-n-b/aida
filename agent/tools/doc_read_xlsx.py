"""
DocReadXlsxTool · 读取 Excel 工作表

读取 .xlsx 文件的指定 sheet，返回 CSV 格式文本或 JSON 数组。
结果可直接喂给 LLM 分析，也可在 step 内结构化消费。
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from .base import Tool, ToolError


class DocReadXlsxTool(Tool):
    @property
    def name(self) -> str:
        return "doc_read_xlsx"

    @property
    def description(self) -> str:
        return (
            "读取 Excel（.xlsx）文件的指定工作表，返回表头 + 行数据。"
            "可按 sheet 名或索引选择，支持限制最大行数。"
            "默认返回 CSV 文本；as_json=true 时返回 JSON 数组（适合程序消费）。"
        )

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "path": {
                    "type": "string",
                    "description": "Excel 文件路径（绝对或相对）",
                    "minLength": 1,
                },
                "sheet": {
                    "type": "string",
                    "description": "工作表名称；省略时读第一个 sheet",
                },
                "max_rows": {
                    "type": "integer",
                    "description": "最多读取的数据行数（不含表头），默认 200",
                    "minimum": 1,
                    "maximum": 5000,
                },
                "as_json": {
                    "type": "boolean",
                    "description": "true 时返回 JSON 数组（每行是 {列名: 值} 字典），默认 false 返回 CSV 文本",
                },
            },
            "required": ["path"],
        }

    def execute(
        self,
        path: str,
        sheet: str = "",
        max_rows: int = 200,
        as_json: bool = False,
        **_: Any,
    ) -> str:
        try:
            import openpyxl
        except ImportError:
            return ToolError("缺少依赖 openpyxl，请运行 pip install openpyxl")

        p = Path(path)
        if not p.is_file():
            return ToolError(f"文件不存在：{path}")
        if p.suffix.lower() not in (".xlsx", ".xlsm", ".xltx", ".xltm"):
            return ToolError(f"不支持的文件格式（需 .xlsx）：{path}")

        try:
            wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        except Exception as e:
            return ToolError(f"无法打开 Excel 文件：{e}")

        # 选 sheet
        ws = None
        if sheet:
            if sheet in wb.sheetnames:
                ws = wb[sheet]
            else:
                available = ", ".join(wb.sheetnames)
                wb.close()
                return ToolError(f"sheet '{sheet}' 不存在，可用：{available}")
        else:
            ws = wb.active

        if ws is None:
            wb.close()
            return ToolError("工作表为空")

        # 读数据
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            return "（工作表为空）"

        headers = [str(c) if c is not None else "" for c in rows[0]]
        data_rows = rows[1 : max_rows + 1]

        if as_json:
            import json
            records = []
            for row in data_rows:
                record = {
                    headers[i]: (row[i] if i < len(row) else None)
                    for i in range(len(headers))
                }
                records.append(record)
            meta = f"// sheet={ws.title} · {len(data_rows)} 行（限 {max_rows}）\n"
            return meta + json.dumps(records, ensure_ascii=False, indent=2, default=str)

        # CSV 格式
        lines = [",".join(f'"{h}"' for h in headers)]
        for row in data_rows:
            cells = []
            for i in range(len(headers)):
                val = row[i] if i < len(row) else ""
                cell_str = "" if val is None else str(val).replace('"', '""')
                cells.append(f'"{cell_str}"')
            lines.append(",".join(cells))

        meta = f"# sheet={ws.title} · {len(data_rows)}/{max(0, len(rows) - 1)} 行（限 {max_rows}）\n"
        return meta + "\n".join(lines)
