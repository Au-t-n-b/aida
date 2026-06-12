"""
BOQ xlsx → Markdown converter (standalone).
Uses officecli for parsing, boq_rules for post-processing.
MD uses embedded HTML tables with inline styles for layout fidelity.

Usage: python xlsx2md.py <source_dir> [output_dir]
"""
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
import openpyxl, os, sys, re, json, subprocess, html as html_mod
from openpyxl.utils import get_column_letter
from boq_rules import post_process_sheet

OFFICECLI_ENV = {**os.environ, "GIT_EXEC_PATH": "D:/app/Git/mingw64/libexec/git-core",
                "OFFICECLI_NO_AUTO_RESIDENT": "1"}


def officecli_parse(filepath, sheet_name, max_row, max_col):
    end_col = get_column_letter(max_col)
    r = subprocess.run(["officecli", "get", filepath, f"{sheet_name}!A1:{end_col}{max_row}", "--json"],
                       capture_output=True, text=True, encoding="utf-8", errors="replace",
                       env=OFFICECLI_ENV, timeout=60)
    if r.returncode != 0: return None
    try: data = json.loads(r.stdout)
    except json.JSONDecodeError: return None
    children = data.get("data", {}).get("results", [{}])[0].get("children", [])
    if not children: return None
    cells = {}
    for cell in children:
        m = re.match(r'/[^/]+/([A-Z]+)(\d+)', cell.get("path", ""))
        if m:
            col_letter, row_num = m.group(1), int(m.group(2))
            col_num = sum((ord(ch) - ord('A') + 1) * (26 ** i) for i, ch in enumerate(reversed(col_letter)))
            cells[(row_num, col_num)] = str(cell.get("text", "") or "")
    if not cells: return None
    return [[cells.get((r, c), "") for c in range(1, max_col + 1)] for r in range(1, max_row + 1)]


def openpyxl_parse(filepath, sheet_name, max_row, max_col):
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb[sheet_name]
        rows = [[ws.cell(row=r, column=c).value for c in range(1, max_col + 1)] for r in range(1, max_row + 1)]
        wb.close(); return rows
    except Exception:
        return None


def _officecli_sheet_meta(filepath):
    r = subprocess.run(["officecli", "get", filepath, "X!A1", "--json"],
        capture_output=True, text=True, encoding="utf-8", errors="replace",
        env=OFFICECLI_ENV, timeout=10)
    m = re.search(r'Available sheets: \[(.*?)\]', (r.stdout or "") + (r.stderr or ""))
    if not m: return None
    sheets = [s.strip() for s in m.group(1).split(",") if s.strip() != "tmp_v"]
    return {sn: {"max_row": 200, "max_col": 30, "hidden_rows": set()} for sn in sheets}

def get_sheet_meta(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception:
        return _officecli_sheet_meta(filepath)
    meta = {}
    for sn in wb.sheetnames:
        if sn == "tmp_v": continue
        ws = wb[sn]
        hidden_rows = {r for r in range(1, ws.max_row + 1)
                       if ws.row_dimensions.get(r) and ws.row_dimensions.get(r).hidden}
        meta[sn] = {"max_row": ws.max_row, "max_col": ws.max_column, "hidden_rows": hidden_rows}
    wb.close(); return meta


def build_md(file_name, source_path, parser_note, sheet_data):
    parts = [f"# {file_name}", "", f"> 文件路径: `{source_path}`", f"> 解析: {parser_note}", "",
             f"> 图例: 黄色高亮行 = 原始Excel隐藏行", ""]

    for sn, rows, hidden_map, first_row in sheet_data:
        parts.append(f"## {sn}")
        parts.append("")
        if not rows:
            parts.append("*(无数据)*\n")
            continue
        parts.append('<div class="table-wrapper"><table>')
        for i, row in enumerate(rows):
            actual_row = first_row + i
            hidden = hidden_map.get(actual_row, False)
            if hidden:
                parts.append('<tr style="background:#fefce8;color:#92400e;border-left:3px solid #f59e0b;font-style:italic">')
                td_attr = ' style="background:#fefce8;color:#92400e;border-color:#fde68a"'
            else:
                parts.append('<tr>')
                td_attr = ''
            for v in row:
                escaped = html_mod.escape(str(v)).replace("\n", "<br>")
                parts.append(f'<td{td_attr}>{escaped}</td>')
            parts.append('</tr>')
        parts.append('</table></div>')
        parts.append("")
    return "\n".join(parts)


def main():
    if len(sys.argv) < 2: print("Usage: python xlsx2md.py <source_dir> [output_dir]"); sys.exit(1)
    source_dir, output_dir = sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else "uniEX-Output"
    project_name = os.path.basename(os.path.abspath(source_dir))
    md_dir = os.path.join(output_dir, project_name, "md")

    files = []
    for root, dirs, filenames in os.walk(source_dir):
        for fn in filenames:
            if fn.endswith(".xlsx") and not fn.startswith("~$"): files.append(os.path.join(root, fn))

    print(f"xlsx2md: {len(files)} files")
    for i, fp in enumerate(sorted(files), 1):
        rel_path = os.path.relpath(fp, source_dir)
        file_name = os.path.splitext(os.path.basename(fp))[0]
        rel_dir = os.path.dirname(rel_path)
        print(f"[{i}/{len(files)}] {os.path.basename(fp)}")

        sheet_meta = get_sheet_meta(fp)
        if sheet_meta is None:
            print(f"  SKIP: cannot discover sheets"); continue
        oc_ok = oc_fail = 0
        sheet_data = []
        for sn, meta in sheet_meta.items():
            raw = officecli_parse(fp, sn, meta["max_row"], meta["max_col"])
            if raw is None: raw = openpyxl_parse(fp, sn, meta["max_row"], meta["max_col"]); oc_fail += 1
            else: oc_ok += 1
            if raw is None: continue
            cleaned, hidden_map, first_row = post_process_sheet(raw, meta["hidden_rows"])
            sheet_data.append((sn, cleaned, hidden_map, first_row))

        content = build_md(file_name, rel_path, f"officecli:{oc_ok} openpyxl:{oc_fail}", sheet_data)
        os.makedirs(os.path.join(md_dir, rel_dir), exist_ok=True)
        mp = os.path.join(md_dir, rel_dir, f"{file_name}.md")
        with open(mp, "w", encoding="utf-8") as f: f.write(content)
        print(f"  -> {os.path.relpath(mp, output_dir)}")
    print(f"\nDone -> {md_dir}/")


if __name__ == "__main__": main()
